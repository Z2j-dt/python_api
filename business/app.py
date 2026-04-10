# -*- coding: utf-8 -*-
"""
业务应用 - 单入口 app.py，合并 sr_api（实时加微监测等）。
business 下仅 app.py，frontend、sql 为资产目录。
"""
import os
import io
import csv
import re as _re
from pathlib import Path
from urllib.parse import quote

from fastapi import FastAPI, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from typing import List, Dict, Any, Optional
import pymysql
from contextlib import contextmanager
from datetime import datetime, date, timedelta

from pydantic import BaseModel

_BUSINESS = Path(__file__).resolve().parent
FRONTEND_DIST = _BUSINESS / "frontend" / "dist"
# 独立运行时 API 在根路径 MOUNT_PATH=""；挂到 run_all 时为 /business/realtime_mv
MOUNT_PATH = (
    (os.environ.get("SR_API_MOUNT_PATH") or "").rstrip("/")
    if os.environ.get("BUSINESS_PORT")
    else (os.environ.get("SR_API_MOUNT_PATH") or "/business/realtime_mv").rstrip("/")
)

# 配置（从 .env 或环境变量）
try:
    from pydantic import BaseSettings
    class _Settings(BaseSettings):
        starrocks_host: str = "127.0.0.1"
        starrocks_port: int = 9030
        starrocks_user: str = "root"
        starrocks_password: str = ""
        starrocks_database: str = "your_database"
        mv_tables: str = "mv_table1,mv_table2"
        class Config:
            env_file = str(_BUSINESS / ".env")
            extra = "ignore"
    _settings = _Settings()
except Exception:
    class _Settings:
        starrocks_host = os.environ.get("STARROCKS_HOST", "127.0.0.1")
        starrocks_port = int(os.environ.get("STARROCKS_PORT", "9030"))
        starrocks_user = os.environ.get("STARROCKS_USER", "root")
        starrocks_password = os.environ.get("STARROCKS_PASSWORD", "")
        starrocks_database = os.environ.get("STARROCKS_DATABASE", "your_database")
        mv_tables = os.environ.get("MV_TABLES", "mv_table1,mv_table2")
    _settings = _Settings()

def _mv_list():
    return [t.strip() for t in _settings.mv_tables.split(",") if t.strip()]

_TEST_PRODUCT_NAME_TOKENS_CN = ("测试", "演示", "沙箱", "预发", "灰度")
_TEST_PRODUCT_NAME_TOKENS_EN = ("test", "uat", "demo", "mock", "staging", "sandbox")


def _is_test_product_name(name: Optional[str]) -> bool:
    s = (name or "").strip()
    if not s:
        return False
    low = s.lower()
    return any(t in s for t in _TEST_PRODUCT_NAME_TOKENS_CN) or any(t in low for t in _TEST_PRODUCT_NAME_TOKENS_EN)


def _get_conn(connect_timeout: int = 10):
    """DB 连接，避免内网不可达时长时间挂死。"""
    return pymysql.connect(
        host=_settings.starrocks_host, port=_settings.starrocks_port,
        user=_settings.starrocks_user, password=_settings.starrocks_password,
        database=_settings.starrocks_database, charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        connect_timeout=connect_timeout,
    )

@contextmanager
def _db_cursor():
    conn = _get_conn()
    try:
        cur = conn.cursor()
        yield cur
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()

# -------------------- 业务配置：只读账号（可选） --------------------
#
# 默认不启用任何限制；如需只读账号，请在服务启动时设置环境变量：
# - BUSINESS_READONLY_USERS: 逗号分隔用户名列表，例如 "yewu_ro,readonly"
#
# 网关/门户侧需要将登录用户名透传到请求头（任一即可）：
# - X-User / X-Username / X-Forwarded-User / X-Portal-User
_READONLY_USERS = {u.strip() for u in (os.environ.get("BUSINESS_READONLY_USERS") or "").split(",") if u.strip()}
_USER_HEADER_KEYS = ("x-user", "x-username", "x-forwarded-user", "x-portal-user")


def _get_request_user(request: Request) -> str:
    # 1) 反向代理/网关注入的用户头
    for k in _USER_HEADER_KEYS:
        v = request.headers.get(k)
        if v:
            return v.strip()
    # 2) 同源门户模式：浏览器会带 portal_auth Cookie（itsdangerous 签名）
    try:
        cookie_val = request.cookies.get("portal_auth") or ""
        secret = os.environ.get("PORTAL_SECRET_KEY") or ""
        if cookie_val and secret:
            try:
                from itsdangerous import URLSafeTimedSerializer
                s = URLSafeTimedSerializer(secret, salt="portal-auth")
                data = s.loads(cookie_val, max_age=int(os.environ.get("PORTAL_AUTH_COOKIE_MAX_AGE", "86400")))
                u = (data.get("u") or "").strip()
                if u:
                    return u
            except Exception:
                pass
    except Exception:
        pass
    # 3) 多端口 iframe：portal_token（短期有效）
    try:
        token = (request.query_params.get("portal_token") or "").strip()
        secret = os.environ.get("PORTAL_SECRET_KEY") or ""
        if token and secret:
            from itsdangerous import URLSafeTimedSerializer
            s = URLSafeTimedSerializer(secret, salt="portal-token")
            data = s.loads(token, max_age=600)
            u = (data.get("u") or "").strip()
            if u:
                return u
    except Exception:
        pass
    return ""


def _reject_if_readonly(request: Request) -> None:
    if not _READONLY_USERS:
        return
    user = _get_request_user(request)
    if user and user in _READONLY_USERS:
        raise HTTPException(status_code=403, detail="只读账号不允许修改配置数据")

# 自营渠道每日汇总表：支持按 dt 日期筛选，默认今天+昨天
DAILY_STATS_TABLE = "mv_scrm_open_channel_tag_stats"

def _query_mv_data(
    table_name: str,
    limit: int = 1000,
    tag_name: Optional[str] = None,
    dt_from: Optional[str] = None,
    dt_to: Optional[str] = None,
    dt_all: bool = False,
) -> List[Dict]:
    if table_name not in _mv_list():
        raise ValueError(f"表 {table_name} 未在配置中")
    with _db_cursor() as cur:
        conditions = []
        args: List[Any] = []
        if tag_name:
            conditions.append("tag_name = %s")
            args.append(tag_name)
        # 自营渠道每日表：按 dt 筛选；未传且非“全部”时默认今天和昨天
        if table_name == DAILY_STATS_TABLE:
            if dt_all:
                pass  # 全部日期，不限制 dt
            elif dt_from and dt_to:
                conditions.append("dt BETWEEN %s AND %s")
                args.extend([dt_from, dt_to])
            else:
                from datetime import date, timedelta
                today = date.today()
                yesterday = today - timedelta(days=1)
                conditions.append("dt BETWEEN %s AND %s")
                args.extend([yesterday.isoformat(), today.isoformat()])
        if conditions:
            where_sql = " AND ".join(conditions)
            args.append(limit)
            order = " ORDER BY dt DESC, open_channel, wechat_customer_tag" if table_name == DAILY_STATS_TABLE else ""
            cur.execute(f"SELECT * FROM `{table_name}` WHERE {where_sql}{order} LIMIT %s", tuple(args))
        else:
            order = " ORDER BY dt DESC" if table_name == DAILY_STATS_TABLE else ""
            cur.execute(f"SELECT * FROM `{table_name}`{order} LIMIT %s", (limit,))
        return cur.fetchall()

def _query_distinct_tags(table_name: str) -> List[str]:
    if table_name not in _mv_list():
        raise ValueError(f"表 {table_name} 未在配置中")
    with _db_cursor() as cur:
        cur.execute(f"SELECT DISTINCT tag_name FROM `{table_name}` WHERE tag_name IS NOT NULL AND tag_name != '' ORDER BY tag_name")
        return [r["tag_name"] for r in cur.fetchall()]

def _get_table_columns(table_name: str) -> List[Dict]:
    with _db_cursor() as cur:
        cur.execute(f"DESCRIBE `{table_name}`")
        return cur.fetchall()


# -------------------- 配置表：开户渠道 & 企微客户标签 --------------------

CONFIG_OPEN_CHANNEL_TABLE = "config_open_channel_tag"
CONFIG_ACTIVITY_CHANNEL_TABLE = "config_activity_channel_tag"
CONFIG_CHANNEL_STAFF_TABLE = "config_channel_staff"
CONFIG_CODE_MAPPING_TABLE = "code_mapping"
CONFIG_STOCK_POSITION_TABLE = "config_stock_position"
CONFIG_OPPORTUNITY_LEAD_TABLE = "config_opportunity_lead"
CONFIG_MORNING_HOT_STOCK_TRACK_TABLE = "config_morning_hot_stock_track"
CONFIG_SALES_ORDER_TABLE = "branch_customer_ext_config"
CONFIG_SIGN_CUSTOMER_GROUP_TABLE = "config_sign_customer_group"

# -------------------- 销售每日进线（活动渠道）：历史事实表（原 mv_scrm_tagged_customers_act） --------------------
MV_SALES_DAILY_LEADS_ACT_TABLE = os.environ.get("SALES_DAILY_LEADS_TABLE", "scrm_tagged_customers_act_hist")
# 月度转化率（简易/复杂均基于同一物化视图；列与 mv_sales_inflow_conversion_monthly.sql 一致）
MV_SALES_INFLOW_CONVERSION_MONTHLY_TABLE = os.environ.get(
    "SALES_INFLOW_CONVERSION_MONTHLY_TABLE", "mv_sales_inflow_conversion_monthly_complex"
)


def _norm_yyyymmdd(s: Optional[str]) -> Optional[str]:
    if s is None:
        return None
    t = str(s).strip()
    if not t:
        return None
    # 兼容 YYYY-MM-DD / YYYYMMDD
    if _re.match(r"^\d{4}-\d{2}-\d{2}$", t):
        return t.replace("-", "")
    if _re.match(r"^\d{8}$", t):
        return t
    return None


class SalesDailyLeadOut(BaseModel):
    name: Optional[str] = None
    user_id: Optional[str] = None
    user_name: Optional[str] = None
    add_time: Optional[str] = None  # yyyymmdd
    tag_name: Optional[str] = None
    group_name: Optional[str] = None
    open_channel: Optional[str] = None


class SalesDailySummaryDailyOut(BaseModel):
    sales_name: Optional[str] = None
    activity_add_cnt: int = 0
    inherit_add_cnt: int = 0
    share_add_cnt: int = 0
    total_add_cnt: int = 0


class SalesDailySummaryMonthlyOut(BaseModel):
    sales_name: Optional[str] = None
    activity_add_cnt: int = 0
    inherit_add_cnt: int = 0
    share_add_cnt: int = 0
    total_add_cnt: int = 0


class SalesInflowConversionMonthlySimpleOut(BaseModel):
    """月度转化率·简易：按月 × 销售一行。"""
    sales_name: Optional[str] = None
    month: Optional[str] = None
    total_add_cnt: int = 0
    cash_order_total: int = 0
    commission_order_total: int = 0
    total_order_cnt: int = 0
    conversion_rate_total: Optional[float] = None


class SalesInflowConversionMonthlyComplexOut(BaseModel):
    """月度转化率·复杂：与物化视图列一致。"""
    sales_name: Optional[str] = None
    month: Optional[str] = None
    total_add_cnt: int = 0
    activity_add_cnt: int = 0
    inherit_add_cnt: int = 0
    share_add_cnt: int = 0
    commission_order_total: int = 0
    commission_order_activity: int = 0
    commission_order_inherit: int = 0
    commission_order_share: int = 0
    cash_order_total: int = 0
    cash_order_activity: int = 0
    cash_order_inherit: int = 0
    cash_order_share: int = 0
    total_order_cnt: int = 0
    order_total_activity: int = 0
    order_total_inherit: int = 0
    order_total_share: int = 0
    conversion_rate_activity: Optional[float] = None
    conversion_rate_inherit: Optional[float] = None
    conversion_rate_share: Optional[float] = None
    conversion_rate_total: Optional[float] = None


class OpenChannelTagBase(BaseModel):
    open_channel: str
    wechat_customer_tag: str


class OpenChannelTagCreate(OpenChannelTagBase):
    pass


class OpenChannelTagUpdate(BaseModel):
    open_channel: Optional[str] = None
    wechat_customer_tag: Optional[str] = None


class OpenChannelTagOut(OpenChannelTagBase):
    id: int
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class ChannelStaffBase(BaseModel):
    branch_name: str  # 营业部
    staff_name: str   # 姓名
    channel_type: str  # 渠道类型（活动渠道/开户渠道）


class ChannelStaffCreate(ChannelStaffBase):
    pass


class ChannelStaffUpdate(BaseModel):
    branch_name: Optional[str] = None
    staff_name: Optional[str] = None
    channel_type: Optional[str] = None


class ChannelStaffOut(ChannelStaffBase):
    id: int
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


class SignCustomerGroupUpdate(BaseModel):
    in_group: int


# -------------------- 配置表：客户中心-商机线索 --------------------

class OpportunityLeadBase(BaseModel):
    biz_category_big: Optional[str] = None   # 业务大类
    biz_category_small: Optional[str] = None # 业务小类
    clue_name: Optional[str] = None          # 线索名称
    is_important: Optional[bool] = None      # 是否重要
    remark: Optional[str] = None             # 备注
    table_name: Optional[str] = None         # 表名


class OpportunityLeadCreate(OpportunityLeadBase):
    pass


class OpportunityLeadUpdate(OpportunityLeadBase):
    pass


class OpportunityLeadOut(OpportunityLeadBase):
    id: int
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


def _row_opportunity_lead_fix(row: Dict[str, Any]) -> Dict[str, Any]:
    if not row:
        return row
    row = _row_config_time_fix(row)
    if "is_important" in row:
        v = row.get("is_important")
        if v is None:
            row["is_important"] = None
        elif isinstance(v, bool):
            row["is_important"] = v
        else:
            try:
                row["is_important"] = bool(int(v))
            except Exception:
                row["is_important"] = bool(v)
    return row


# -------------------- 配置表：客户中心-早盘人气股战绩追踪 --------------------

class MorningHotStockTrackBase(BaseModel):
    tg_name: Optional[str] = None      # 投顾/老师名称
    biz_date: Optional[str] = None     # YYYY-MM-DD
    stock_name: Optional[str] = None   # 人气股
    stock_code: Optional[str] = None   # 代码
    remark: Optional[str] = None       # 备注


class MorningHotStockTrackCreate(MorningHotStockTrackBase):
    pass


class MorningHotStockTrackUpdate(MorningHotStockTrackBase):
    pass


class MorningHotStockTrackOut(MorningHotStockTrackBase):
    id: int
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


def _row_morning_hot_stock_track_fix(row: Dict[str, Any]) -> Dict[str, Any]:
    if not row:
        return row
    row = dict(row)
    # 时间字段
    row = _row_config_time_fix(row)
    # 日期字段
    if "biz_date" in row and row.get("biz_date") is not None:
        row["biz_date"] = _parse_date(row.get("biz_date"))
    return row


_DT_FMT = "%Y-%m-%d %H:%M:%S"


def _fmt_dt(v: Any) -> Optional[str]:
    """把数据库的时间字段格式化为 'YYYY-MM-DD HH:MM:SS'。"""
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime(_DT_FMT)
    # 兼容某些驱动/查询返回 str
    try:
        s = str(v)
        if "T" in s and len(s) >= 19:
            # 2026-02-06T13:13:13 -> 2026-02-06 13:13:13
            s = s.replace("T", " ", 1)
        return s
    except Exception:
        return None


def _row_config_time_fix(row: Dict[str, Any]) -> Dict[str, Any]:
    if not row:
        return row
    row = dict(row)
    if "created_at" in row:
        row["created_at"] = _fmt_dt(row.get("created_at"))
    if "updated_at" in row:
        row["updated_at"] = _fmt_dt(row.get("updated_at"))
    return row


def _row_code_mapping_fix(row: Dict[str, Any]) -> Dict[str, Any]:
    """格式化 code_mapping 的时间/数值字段，便于前端展示。"""
    if not row:
        return row
    row = dict(row)
    if "created_time" in row:
        row["created_time"] = _fmt_dt(row.get("created_time"))
    # DECIMAL 可能返回 Decimal，这里转成 float（前端只做展示/编辑）
    if "stat_cost" in row and row.get("stat_cost") is not None:
        try:
            row["stat_cost"] = float(row["stat_cost"])
        except Exception:
            pass
    return row


class CodeMappingBase(BaseModel):
    code_value: str
    description: Optional[str] = None
    stat_cost: Optional[float] = None
    channel_name: Optional[str] = None


class CodeMappingCreate(CodeMappingBase):
    id: int


class CodeMappingUpdate(BaseModel):
    code_value: Optional[str] = None
    description: Optional[str] = None
    stat_cost: Optional[float] = None
    channel_name: Optional[str] = None


class CodeMappingOut(CodeMappingBase):
    id: int
    created_time: Optional[str] = None


# -------------------- 配置表：股票仓位/买卖 --------------------

class StockPositionBase(BaseModel):
    product_name: Optional[str] = None
    trade_date: Optional[str] = None   # YYYY-MM-DD
    row_id: Optional[int] = None       # 序号（同一产品+日期内从 1 开始递增）
    stock_code: Optional[str] = None
    stock_name: Optional[str] = None
    position_pct: Optional[float] = None
    side: Optional[str] = None        # 买入 / 卖出
    price: Optional[float] = None


class StockPositionCreate(StockPositionBase):
    stock_code: str
    position_pct: float
    side: str
    price: float


class StockPositionUpdate(BaseModel):
    product_name: Optional[str] = None
    trade_date: Optional[str] = None
    stock_code: Optional[str] = None
    stock_name: Optional[str] = None
    position_pct: Optional[float] = None
    side: Optional[str] = None
    price: Optional[float] = None


class StockPositionOut(StockPositionBase):
    id: Optional[int] = None
    created_at: Optional[str] = None
    updated_at: Optional[str] = None


def _row_stock_position_fix(row: Dict[str, Any]) -> Dict[str, Any]:
    if not row:
        return row
    row = dict(row)
    if "trade_date" in row and row.get("trade_date") is not None:
        v = row["trade_date"]
        if hasattr(v, "strftime"):
            row["trade_date"] = v.strftime("%Y-%m-%d")
        else:
            row["trade_date"] = str(v)[:10]
    if "created_at" in row:
        row["created_at"] = _fmt_dt(row.get("created_at"))
    if "updated_at" in row:
        row["updated_at"] = _fmt_dt(row.get("updated_at"))
    if "row_id" in row and row.get("row_id") is not None:
        try:
            row["row_id"] = int(row["row_id"])
        except Exception:
            pass
    for key in ("position_pct", "price"):
        if key in row and row.get(key) is not None:
            try:
                row[key] = float(row[key])
            except Exception:
                pass
    return row


def _next_config_id(cur, table_name: str) -> int:
    """
    返回最小可用正整数 id（从 1 开始，删除后会补空缺）。
    适用于配置表这种小表。
    """
    # 之前这里漏了分支：当 `id=1` 已存在时会返回 None，
    # 最终插入时触发 StarRocks 的 "NULL value in non-nullable column 'id'"。
    cur.execute(f"SELECT id FROM `{table_name}` ORDER BY id ASC")
    rows = cur.fetchall() or []
    ids: List[int] = []
    for r in rows:
        v = (r or {}).get("id")
        if v is None:
            continue
        try:
            ids.append(int(v))
        except Exception:
            continue

    # 找最小缺失的正整数，从 1 开始递增
    next_id = 1
    for v in ids:
        if v < next_id:
            continue
        if v == next_id:
            next_id += 1
            continue
        # v > next_id：说明 next_id 这号是缺的
        break
    return next_id if next_id >= 1 else 1


def _parse_month_yyyy_mm(v: Optional[str]) -> str:
    """解析 YYYY-MM，非法时返回当前月。"""
    s = (v or "").strip()
    if s:
        try:
            return datetime.strptime(s, "%Y-%m").strftime("%Y-%m")
        except Exception:
            pass
    return datetime.now().strftime("%Y-%m")


def _next_row_id_for_date(cur, product_name: str, trade_date: str) -> int:
    """
    返回同一产品+日期下下一个可用 row_id（从 1 开始递增）。
    """
    sql = (
        f"SELECT MAX(row_id) AS max_id FROM `{CONFIG_STOCK_POSITION_TABLE}` "
        f"WHERE product_name <=> %s AND trade_date = %s"
    )
    cur.execute(sql, (product_name or None, trade_date))
    row = cur.fetchone() or {}
    max_id = row.get("max_id")
    try:
        val = int(max_id)
    except Exception:
        val = 0
    return val + 1 if val >= 0 else 1

app = FastAPI(title="StarRocks MV 数据服务", description="查询 StarRocks 物化视图", version="1.0.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


@app.middleware("http")
async def allow_frame_embedding(request, call_next):
    """允许被门户 iframe 嵌入（门户与业务不同端口，需允许跨端口嵌入）"""
    response = await call_next(request)
    response.headers["Content-Security-Policy"] = "frame-ancestors *"
    return response


@app.get("/health")
async def health():
    return {"status": "ok", "message": "StarRocks MV 数据服务运行中"}

@app.get("/api/tables")
async def get_tables() -> List[str]:
    return _mv_list()

@app.get("/api/data/{table_name}")
async def get_table_data(
    table_name: str,
    limit: int = 1000,
    tag_name: Optional[str] = None,
    dt_from: Optional[str] = None,
    dt_to: Optional[str] = None,
    dt_all: Optional[bool] = None,
) -> Dict[str, Any]:
    if table_name not in _mv_list():
        raise HTTPException(status_code=404, detail=f"表 {table_name} 未配置或不存在")
    try:
        data = _query_mv_data(
            table_name,
            limit=limit,
            tag_name=tag_name or None,
            dt_from=dt_from,
            dt_to=dt_to,
            dt_all=bool(dt_all) if dt_all is not None else False,
        )
        return {"table": table_name, "count": len(data), "data": data}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/tags/{table_name}")
async def get_tags(table_name: str) -> List[str]:
    if table_name not in _mv_list():
        raise HTTPException(status_code=404, detail=f"表 {table_name} 未配置或不存在")
    try:
        return _query_distinct_tags(table_name)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/schema/{table_name}")
async def get_table_schema(table_name: str) -> Dict[str, Any]:
    if table_name not in _mv_list():
        raise HTTPException(status_code=404, detail=f"表 {table_name} 未配置或不存在")
    try:
        return {"table": table_name, "columns": _get_table_columns(table_name)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：开户渠道 & 企微客户标签配置 --------------------

@app.get("/api/config/open-channel-tags", response_model=List[OpenChannelTagOut])
async def list_open_channel_tags() -> List[OpenChannelTagOut]:
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_OPEN_CHANNEL_TABLE}` ORDER BY id ASC"
            )
            rows = cur.fetchall()
        return [OpenChannelTagOut(**_row_config_time_fix(row)) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/open-channel-tags", response_model=OpenChannelTagOut)
async def create_open_channel_tag(body: OpenChannelTagCreate, request: Request) -> OpenChannelTagOut:
    _reject_if_readonly(request)
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            next_id = _next_config_id(cur, CONFIG_OPEN_CHANNEL_TABLE)
            cur.execute(
                f"INSERT INTO `{CONFIG_OPEN_CHANNEL_TABLE}` "
                f"(id, open_channel, wechat_customer_tag, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s)",
                (next_id, body.open_channel, body.wechat_customer_tag, now_str, now_str),
            )
            # 重新查询最新一条记录返回
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_OPEN_CHANNEL_TABLE}` WHERE id = %s",
                (next_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return OpenChannelTagOut(**_row_config_time_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/open-channel-tags/{item_id}", response_model=OpenChannelTagOut)
async def update_open_channel_tag(item_id: int, body: OpenChannelTagUpdate, request: Request) -> OpenChannelTagOut:
    _reject_if_readonly(request)
    if body.open_channel is None and body.wechat_customer_tag is None:
        raise HTTPException(status_code=400, detail="至少提供一个需要更新的字段")
    try:
        # 兼容 StarRocks：部分表不支持 UPDATE（会报 does not support update）
        # 用“查旧值 -> DELETE -> INSERT”实现修改，保留 created_at，仅更新 updated_at。
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_OPEN_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            old = cur.fetchone()
            if not old:
                raise HTTPException(status_code=404, detail="记录不存在")

            new_open = body.open_channel if body.open_channel is not None else old.get("open_channel")
            new_tag = body.wechat_customer_tag if body.wechat_customer_tag is not None else old.get("wechat_customer_tag")
            created_at = _fmt_dt(old.get("created_at")) or now_str

            cur.execute(
                f"DELETE FROM `{CONFIG_OPEN_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            cur.execute(
                f"INSERT INTO `{CONFIG_OPEN_CHANNEL_TABLE}` "
                f"(id, open_channel, wechat_customer_tag, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s)",
                (item_id, new_open, new_tag, created_at, now_str),
            )
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_OPEN_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return OpenChannelTagOut(**_row_config_time_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/open-channel-tags/{item_id}")
async def delete_open_channel_tag(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"DELETE FROM `{CONFIG_OPEN_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：活动渠道字典配置 --------------------

@app.get("/api/config/activity-channel-tags", response_model=List[OpenChannelTagOut])
async def list_activity_channel_tags() -> List[OpenChannelTagOut]:
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_ACTIVITY_CHANNEL_TABLE}` ORDER BY id ASC"
            )
            rows = cur.fetchall()
        return [OpenChannelTagOut(**_row_config_time_fix(row)) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/activity-channel-tags", response_model=OpenChannelTagOut)
async def create_activity_channel_tag(body: OpenChannelTagCreate, request: Request) -> OpenChannelTagOut:
    _reject_if_readonly(request)
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            next_id = _next_config_id(cur, CONFIG_ACTIVITY_CHANNEL_TABLE)
            cur.execute(
                f"INSERT INTO `{CONFIG_ACTIVITY_CHANNEL_TABLE}` "
                f"(id, open_channel, wechat_customer_tag, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s)",
                (next_id, body.open_channel, body.wechat_customer_tag, now_str, now_str),
            )
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_ACTIVITY_CHANNEL_TABLE}` WHERE id = %s",
                (next_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return OpenChannelTagOut(**_row_config_time_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/activity-channel-tags/{item_id}", response_model=OpenChannelTagOut)
async def update_activity_channel_tag(item_id: int, body: OpenChannelTagUpdate, request: Request) -> OpenChannelTagOut:
    _reject_if_readonly(request)
    if body.open_channel is None and body.wechat_customer_tag is None:
        raise HTTPException(status_code=400, detail="至少提供一个需要更新的字段")
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_ACTIVITY_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            old = cur.fetchone()
            if not old:
                raise HTTPException(status_code=404, detail="记录不存在")

            new_open = body.open_channel if body.open_channel is not None else old.get("open_channel")
            new_tag = body.wechat_customer_tag if body.wechat_customer_tag is not None else old.get("wechat_customer_tag")
            created_at = _fmt_dt(old.get("created_at")) or now_str

            cur.execute(
                f"DELETE FROM `{CONFIG_ACTIVITY_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            cur.execute(
                f"INSERT INTO `{CONFIG_ACTIVITY_CHANNEL_TABLE}` "
                f"(id, open_channel, wechat_customer_tag, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s)",
                (item_id, new_open, new_tag, created_at, now_str),
            )
            cur.execute(
                f"SELECT id, open_channel, wechat_customer_tag, created_at, updated_at "
                f"FROM `{CONFIG_ACTIVITY_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return OpenChannelTagOut(**_row_config_time_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/activity-channel-tags/{item_id}")
async def delete_activity_channel_tag(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"DELETE FROM `{CONFIG_ACTIVITY_CHANNEL_TABLE}` WHERE id = %s",
                (item_id,),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：销售每日进线数据表（活动渠道） --------------------

@app.get("/api/sales-daily-leads/latest-date")
async def sales_daily_leads_latest_date() -> Dict[str, Any]:
    try:
        with _db_cursor() as cur:
            cur.execute(f"SELECT MAX(add_time) AS d FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`")
            row = cur.fetchone() or {}
        d = _norm_yyyymmdd(row.get("d"))
        return {"date": d or ""}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sales-daily-leads", response_model=List[SalesDailyLeadOut])
async def sales_daily_leads_list(date: Optional[str] = None) -> List[SalesDailyLeadOut]:
    """
    默认返回最新一天；传 date 支持 YYYY-MM-DD 或 YYYYMMDD（按 add_time=yyyymmdd 过滤）。
    """
    try:
        target = _norm_yyyymmdd(date)
        if not target:
            with _db_cursor() as cur:
                cur.execute(f"SELECT MAX(add_time) AS d FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`")
                row = cur.fetchone() or {}
                target = _norm_yyyymmdd(row.get("d"))
        if not target:
            return []

        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  name,
                  CAST(user_id AS STRING) AS user_id,
                  user_name,
                  add_time,
                  tag_name,
                  group_name,
                  open_channel
                FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`
                WHERE add_time = %s
                ORDER BY user_name ASC, name ASC
                """,
                (target,),
            )
            rows = cur.fetchall() or []
        return [SalesDailyLeadOut(**dict(r)) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


def _normalize_sales_name(name: Any) -> str:
    """
    销售名称归并：
    - 张三A2 / 张三a2 -> 张三
    - 保留无法识别后缀的原名
    """
    s = str(name or "").strip()
    if not s:
        return "-"
    # 口径改为“中文字符相同即同一销售”：提取全部中文字符作为归并键
    zh = "".join(_re.findall(r"[\u4e00-\u9fff]+", s))
    if zh:
        return zh
    # 兜底：若无中文，再尝试移除尾部字母数字后缀
    s = _re.sub(r"[A-Za-z]\d+$", "", s).strip()
    return s or "-"


def _as_int(v: Any, default: int = 0) -> int:
    try:
        if v is None:
            return default
        return int(v)
    except Exception:
        return default


def _as_float_opt(v: Any) -> Optional[float]:
    if v is None:
        return None
    try:
        return float(v)
    except Exception:
        return None


def _fmt_pct(v: Any, digits: int = 2) -> str:
    """
    将小数（如 0.0123）格式化为百分比字符串（如 1.23%）。
    """
    f = _as_float_opt(v)
    if f is None:
        return ""
    try:
        return f"{(f * 100):.{int(digits)}f}%"
    except Exception:
        return ""


@app.get("/api/sales-daily-leads/summary/daily", response_model=List[SalesDailySummaryDailyOut])
async def sales_daily_leads_summary_daily(date: Optional[str] = None) -> List[SalesDailySummaryDailyOut]:
    """
    日度汇总：
    - 默认最新一天
    - 维度：销售名(按中文字符归并)
    - 指标：活动/继承/共享 三类加微数
    - 最后一行追加三类汇总
    """
    try:
        target = _norm_yyyymmdd(date)
        if not target:
            with _db_cursor() as cur:
                cur.execute(f"SELECT MAX(add_time) AS d FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`")
                row = cur.fetchone() or {}
                target = _norm_yyyymmdd(row.get("d"))
        if not target:
            return []

        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  COALESCE(NULLIF(TRIM(CAST(user_name AS STRING)), ''), '-') AS sales_name,
                  COALESCE(NULLIF(TRIM(CAST(open_channel AS STRING)), ''), '-') AS open_channel,
                  COUNT(*) AS add_cnt
                FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`
                WHERE add_time = %s
                GROUP BY
                  COALESCE(NULLIF(TRIM(CAST(user_name AS STRING)), ''), '-'),
                  COALESCE(NULLIF(TRIM(CAST(open_channel AS STRING)), ''), '-')
                """,
                (target,),
            )
            rows = cur.fetchall() or []

        merged: Dict[str, Dict[str, int]] = {}
        total_activity = 0
        total_inherit = 0
        total_share = 0
        for r in rows:
            d = dict(r)
            sales_name = _normalize_sales_name(d.get("sales_name"))
            open_channel = str(d.get("open_channel") or "").strip()
            try:
                add_cnt = int(d.get("add_cnt") or 0)
            except Exception:
                add_cnt = 0
            if sales_name not in merged:
                merged[sales_name] = {"activity_add_cnt": 0, "inherit_add_cnt": 0, "share_add_cnt": 0}

            if "活动" in open_channel:
                merged[sales_name]["activity_add_cnt"] += add_cnt
                total_activity += add_cnt
            elif "继承" in open_channel:
                merged[sales_name]["inherit_add_cnt"] += add_cnt
                total_inherit += add_cnt
            elif "共享" in open_channel:
                merged[sales_name]["share_add_cnt"] += add_cnt
                total_share += add_cnt

        items: List[SalesDailySummaryDailyOut] = [
            SalesDailySummaryDailyOut(
                sales_name=sales_name,
                activity_add_cnt=cnts["activity_add_cnt"],
                inherit_add_cnt=cnts["inherit_add_cnt"],
                share_add_cnt=cnts["share_add_cnt"],
                total_add_cnt=cnts["activity_add_cnt"] + cnts["inherit_add_cnt"] + cnts["share_add_cnt"],
            )
            for sales_name, cnts in merged.items()
        ]
        items.sort(key=lambda x: (x.sales_name or ""))
        items.append(
            SalesDailySummaryDailyOut(
                sales_name="汇总",
                activity_add_cnt=total_activity,
                inherit_add_cnt=total_inherit,
                share_add_cnt=total_share,
                total_add_cnt=total_activity + total_inherit + total_share,
            )
        )
        return items
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sales-daily-leads/summary/monthly", response_model=List[SalesDailySummaryMonthlyOut])
async def sales_daily_leads_summary_monthly(month: Optional[str] = None) -> List[SalesDailySummaryMonthlyOut]:
    """
    月度汇总：按销售(中文归并)统计活动/继承/共享，默认当前月份。
    """
    m = _parse_month_yyyy_mm(month) or datetime.now().strftime("%Y-%m")
    like_prefix = m.replace("-", "")
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  COALESCE(NULLIF(TRIM(CAST(user_name AS STRING)), ''), '-') AS sales_name,
                  COALESCE(NULLIF(TRIM(CAST(open_channel AS STRING)), ''), '-') AS open_channel,
                  COUNT(*) AS add_cnt
                FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`
                WHERE add_time LIKE %s
                GROUP BY
                  COALESCE(NULLIF(TRIM(CAST(user_name AS STRING)), ''), '-'),
                  COALESCE(NULLIF(TRIM(CAST(open_channel AS STRING)), ''), '-')
                """,
                (f"{like_prefix}%",),
            )
            rows = cur.fetchall() or []
        merged: Dict[str, Dict[str, int]] = {}
        total_activity = 0
        total_inherit = 0
        total_share = 0
        for r in rows:
            d = dict(r)
            sales_name = _normalize_sales_name(d.get("sales_name"))
            open_channel = str(d.get("open_channel") or "").strip()
            try:
                add_cnt = int(d.get("add_cnt") or 0)
            except Exception:
                add_cnt = 0
            if sales_name not in merged:
                merged[sales_name] = {"activity_add_cnt": 0, "inherit_add_cnt": 0, "share_add_cnt": 0}
            if "活动" in open_channel:
                merged[sales_name]["activity_add_cnt"] += add_cnt
                total_activity += add_cnt
            elif "继承" in open_channel:
                merged[sales_name]["inherit_add_cnt"] += add_cnt
                total_inherit += add_cnt
            elif "共享" in open_channel:
                merged[sales_name]["share_add_cnt"] += add_cnt
                total_share += add_cnt

        items: List[SalesDailySummaryMonthlyOut] = [
            SalesDailySummaryMonthlyOut(
                sales_name=sales_name,
                activity_add_cnt=cnts["activity_add_cnt"],
                inherit_add_cnt=cnts["inherit_add_cnt"],
                share_add_cnt=cnts["share_add_cnt"],
                total_add_cnt=cnts["activity_add_cnt"] + cnts["inherit_add_cnt"] + cnts["share_add_cnt"],
            )
            for sales_name, cnts in merged.items()
        ]
        items.sort(key=lambda x: (x.sales_name or ""))
        items.append(
            SalesDailySummaryMonthlyOut(
                sales_name="汇总",
                activity_add_cnt=total_activity,
                inherit_add_cnt=total_inherit,
                share_add_cnt=total_share,
                total_add_cnt=total_activity + total_inherit + total_share,
            )
        )
        return items
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sales-daily-leads/summary/monthly/export.csv")
async def sales_daily_leads_summary_monthly_export_csv(month: Optional[str] = None, all: Optional[str] = None):
    """
    月度汇总下载：固定导出全量（不按月份筛选）。
    """
    try:
        rows: List[Dict[str, Any]] = []
        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  SUBSTRING(CAST(add_time AS STRING), 1, 6) AS ym,
                  COALESCE(NULLIF(TRIM(CAST(open_channel AS STRING)), ''), '-') AS open_channel,
                  COUNT(*) AS add_cnt
                FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`
                WHERE add_time IS NOT NULL
                  AND LENGTH(CAST(add_time AS STRING)) >= 6
                GROUP BY
                  SUBSTRING(CAST(add_time AS STRING), 1, 6),
                  COALESCE(NULLIF(TRIM(CAST(open_channel AS STRING)), ''), '-')
                ORDER BY ym DESC, add_cnt DESC, open_channel ASC
                """
            )
            rows = cur.fetchall() or []

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=["月份", "渠道", "加微人数"])
        writer.writeheader()
        for r in rows:
            ym_raw = str(r.get("ym") or "").strip()
            ym = f"{ym_raw[:4]}-{ym_raw[4:6]}" if len(ym_raw) >= 6 else ym_raw
            writer.writerow(
                {
                    "月份": ym,
                    "渠道": r.get("open_channel") or "-",
                    "加微人数": int(r.get("add_cnt") or 0),
                }
            )

        data = output.getvalue().encode("utf-8-sig")
        filename = quote("sales_daily_leads_monthly_summary_all.csv")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出CSV失败: {e}")


@app.get("/api/sales-daily-leads/export.csv")
async def sales_daily_leads_export_csv(dt_from: Optional[str] = None, dt_to: Optional[str] = None):
    """
    按 add_time(yyyymmdd) 导出区间数据。dt_from/dt_to 支持 YYYY-MM-DD 或 YYYYMMDD。
    """
    try:
        f = _norm_yyyymmdd(dt_from)
        t = _norm_yyyymmdd(dt_to)
        if not f or not t:
            raise HTTPException(status_code=400, detail="dt_from/dt_to 必填，格式支持 YYYY-MM-DD 或 YYYYMMDD")
        if f > t:
            raise HTTPException(status_code=400, detail="dt_from 不能大于 dt_to")

        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  name,
                  CAST(user_id AS STRING) AS user_id,
                  user_name,
                  add_time,
                  tag_name,
                  group_name,
                  open_channel
                FROM `{MV_SALES_DAILY_LEADS_ACT_TABLE}`
                WHERE add_time BETWEEN %s AND %s
                ORDER BY add_time DESC, user_name ASC, name ASC
                """,
                (f, t),
            )
            rows = cur.fetchall() or []

        cols = [
            ("微信昵称", "name"),
            ("员工id", "user_id"),
            ("员工名称", "user_name"),
            ("添加时间", "add_time"),
            ("标签名称", "tag_name"),
            ("标签组名称", "group_name"),
            ("渠道", "open_channel"),
        ]

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[c[0] for c in cols])
        writer.writeheader()
        for r in rows:
            d = dict(r)
            out: Dict[str, Any] = {}
            for cn, en in cols:
                out[cn] = d.get(en)
            writer.writerow(out)

        data = output.getvalue().encode("utf-8-sig")
        filename = quote(f"sales_daily_leads_{f}_{t}.csv")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出CSV失败: {e}")


@app.get("/api/sales-daily-leads/conversion/monthly")
async def sales_daily_leads_conversion_monthly(
    month: Optional[str] = None,
    mode: str = "simple",
) -> List[Any]:
    """
    月度转化率（物化视图）：默认当前月；
    - mode=simple：简易列
    - mode=complex：全量列
    - mode=unopened：未开户（仅 销售/总加微数/订单数/转化率）
    """
    m = _parse_month_yyyy_mm(month)
    md = (mode or "simple").strip().lower()
    if md not in ("simple", "complex", "unopened"):
        raise HTTPException(status_code=400, detail="mode 须为 simple / complex / unopened")
    tbl = MV_SALES_INFLOW_CONVERSION_MONTHLY_TABLE
    try:
        if md == "simple":
            with _db_cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      sales_name,
                      `month`,
                      total_add_cnt,
                      cash_order_total,
                      commission_order_total,
                      total_order_cnt,
                      conversion_rate_total
                    FROM `{tbl}`
                    WHERE `month` = %s
                    ORDER BY sales_name ASC
                    """,
                    (m,),
                )
                rows = cur.fetchall() or []
            return [
                SalesInflowConversionMonthlySimpleOut(
                    sales_name=d.get("sales_name"),
                    month=d.get("month"),
                    total_add_cnt=_as_int(d.get("total_add_cnt")),
                    cash_order_total=_as_int(d.get("cash_order_total")),
                    commission_order_total=_as_int(d.get("commission_order_total")),
                    total_order_cnt=_as_int(d.get("total_order_cnt")),
                    conversion_rate_total=_as_float_opt(d.get("conversion_rate_total")),
                )
                for d in (dict(r) for r in rows)
            ]
        if md == "unopened":
            with _db_cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      sales_name,
                      `month`,
                      unopened_add_cnt AS total_add_cnt,
                      order_total_unopened AS total_order_cnt,
                      conversion_rate_unopened AS conversion_rate_total
                    FROM `{tbl}`
                    WHERE `month` = %s
                      AND COALESCE(unopened_add_cnt, 0) > 0
                    ORDER BY sales_name ASC
                    """,
                    (m,),
                )
                rows = cur.fetchall() or []
            return [
                SalesInflowConversionMonthlySimpleOut(
                    sales_name=d.get("sales_name"),
                    month=d.get("month"),
                    total_add_cnt=_as_int(d.get("total_add_cnt")),
                    cash_order_total=0,
                    commission_order_total=0,
                    total_order_cnt=_as_int(d.get("total_order_cnt")),
                    conversion_rate_total=_as_float_opt(d.get("conversion_rate_total")),
                )
                for d in (dict(r) for r in rows)
            ]
        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT
                  sales_name,
                  `month`,
                  total_add_cnt,
                  activity_add_cnt,
                  inherit_add_cnt,
                  share_add_cnt,
                  commission_order_total,
                  commission_order_activity,
                  commission_order_inherit,
                  commission_order_share,
                  cash_order_total,
                  cash_order_activity,
                  cash_order_inherit,
                  cash_order_share,
                  total_order_cnt,
                  order_total_activity,
                  order_total_inherit,
                  order_total_share,
                  conversion_rate_activity,
                  conversion_rate_inherit,
                  conversion_rate_share,
                  conversion_rate_total
                FROM `{tbl}`
                WHERE `month` = %s
                ORDER BY sales_name ASC
                """,
                (m,),
            )
            rows = cur.fetchall() or []
        return [
            SalesInflowConversionMonthlyComplexOut(
                sales_name=d.get("sales_name"),
                month=d.get("month"),
                total_add_cnt=_as_int(d.get("total_add_cnt")),
                activity_add_cnt=_as_int(d.get("activity_add_cnt")),
                inherit_add_cnt=_as_int(d.get("inherit_add_cnt")),
                share_add_cnt=_as_int(d.get("share_add_cnt")),
                commission_order_total=_as_int(d.get("commission_order_total")),
                commission_order_activity=_as_int(d.get("commission_order_activity")),
                commission_order_inherit=_as_int(d.get("commission_order_inherit")),
                commission_order_share=_as_int(d.get("commission_order_share")),
                cash_order_total=_as_int(d.get("cash_order_total")),
                cash_order_activity=_as_int(d.get("cash_order_activity")),
                cash_order_inherit=_as_int(d.get("cash_order_inherit")),
                cash_order_share=_as_int(d.get("cash_order_share")),
                total_order_cnt=_as_int(d.get("total_order_cnt")),
                order_total_activity=_as_int(d.get("order_total_activity")),
                order_total_inherit=_as_int(d.get("order_total_inherit")),
                order_total_share=_as_int(d.get("order_total_share")),
                conversion_rate_activity=_as_float_opt(d.get("conversion_rate_activity")),
                conversion_rate_inherit=_as_float_opt(d.get("conversion_rate_inherit")),
                conversion_rate_share=_as_float_opt(d.get("conversion_rate_share")),
                conversion_rate_total=_as_float_opt(d.get("conversion_rate_total")),
            )
            for d in (dict(r) for r in rows)
        ]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/sales-daily-leads/conversion/monthly/export.csv")
async def sales_daily_leads_conversion_monthly_export_csv(mode: str = "simple"):
    """
    月度转化率 CSV 全量导出（不按月份过滤，与列表筛选独立）。
    """
    md = (mode or "simple").strip().lower()
    if md not in ("simple", "complex", "unopened"):
        raise HTTPException(status_code=400, detail="mode 须为 simple / complex / unopened")
    tbl = MV_SALES_INFLOW_CONVERSION_MONTHLY_TABLE
    try:
        if md == "unopened":
            with _db_cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      sales_name,
                      `month`,
                      unopened_add_cnt AS total_add_cnt,
                      order_total_unopened AS total_order_cnt,
                      conversion_rate_unopened AS conversion_rate_total
                    FROM `{tbl}`
                    WHERE COALESCE(unopened_add_cnt, 0) > 0
                    ORDER BY `month` DESC, sales_name ASC
                    """
                )
                rows = cur.fetchall() or []
            fieldnames = ["销售", "月份", "总加微数", "订单数", "转化率"]
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                d = dict(r)
                writer.writerow(
                    {
                        "销售": d.get("sales_name") or "",
                        "月份": d.get("month") or "",
                        "总加微数": _as_int(d.get("total_add_cnt")),
                        "订单数": _as_int(d.get("total_order_cnt")),
                        "转化率": _fmt_pct(d.get("conversion_rate_total"), 2),
                    }
                )
            data = output.getvalue().encode("utf-8-sig")
            filename = quote("sales_inflow_conversion_monthly_unopened_all.csv")
            return StreamingResponse(
                io.BytesIO(data),
                media_type="text/csv; charset=utf-8",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
        if md == "simple":
            with _db_cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      sales_name,
                      `month`,
                      total_add_cnt,
                      cash_order_total,
                      commission_order_total,
                      total_order_cnt,
                      conversion_rate_total
                    FROM `{tbl}`
                    ORDER BY `month` DESC, sales_name ASC
                    """
                )
                rows = cur.fetchall() or []
            fieldnames = ["销售", "月份", "总加微数", "现金订单数", "升佣订单数", "总计订单数", "转化率"]
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                d = dict(r)
                writer.writerow(
                    {
                        "销售": d.get("sales_name") or "",
                        "月份": d.get("month") or "",
                        "总加微数": _as_int(d.get("total_add_cnt")),
                        "现金订单数": _as_int(d.get("cash_order_total")),
                        "升佣订单数": _as_int(d.get("commission_order_total")),
                        "总计订单数": _as_int(d.get("total_order_cnt")),
                        "转化率": _fmt_pct(d.get("conversion_rate_total"), 2),
                    }
                )
        else:
            with _db_cursor() as cur:
                cur.execute(
                    f"""
                    SELECT
                      sales_name,
                      `month`,
                      total_add_cnt,
                      activity_add_cnt,
                      inherit_add_cnt,
                      share_add_cnt,
                      commission_order_total,
                      commission_order_activity,
                      commission_order_inherit,
                      commission_order_share,
                      cash_order_total,
                      cash_order_activity,
                      cash_order_inherit,
                      cash_order_share,
                      total_order_cnt,
                      order_total_activity,
                      order_total_inherit,
                      order_total_share,
                      conversion_rate_activity,
                      conversion_rate_inherit,
                      conversion_rate_share,
                      conversion_rate_total
                    FROM `{tbl}`
                    ORDER BY `month` DESC, sales_name ASC
                    """
                )
                rows = cur.fetchall() or []
            fieldnames = [
                "销售",
                "月份",
                "总加微数",
                "活动加微",
                "继承加微",
                "共享加微",
                "升佣总订单",
                "升佣活动",
                "升佣继承",
                "升佣共享",
                "现金总订单",
                "现金活动",
                "现金继承",
                "现金共享",
                "总订单数",
                "订单活动",
                "订单继承",
                "订单共享",
                "转化率活动",
                "转化率继承",
                "转化率共享",
                "转化率总计",
            ]
            output = io.StringIO()
            writer = csv.DictWriter(output, fieldnames=fieldnames)
            writer.writeheader()
            for r in rows:
                d = dict(r)
                writer.writerow(
                    {
                        "销售": d.get("sales_name") or "",
                        "月份": d.get("month") or "",
                        "总加微数": _as_int(d.get("total_add_cnt")),
                        "活动加微": _as_int(d.get("activity_add_cnt")),
                        "继承加微": _as_int(d.get("inherit_add_cnt")),
                        "共享加微": _as_int(d.get("share_add_cnt")),
                        "升佣总订单": _as_int(d.get("commission_order_total")),
                        "升佣活动": _as_int(d.get("commission_order_activity")),
                        "升佣继承": _as_int(d.get("commission_order_inherit")),
                        "升佣共享": _as_int(d.get("commission_order_share")),
                        "现金总订单": _as_int(d.get("cash_order_total")),
                        "现金活动": _as_int(d.get("cash_order_activity")),
                        "现金继承": _as_int(d.get("cash_order_inherit")),
                        "现金共享": _as_int(d.get("cash_order_share")),
                        "总订单数": _as_int(d.get("total_order_cnt")),
                        "订单活动": _as_int(d.get("order_total_activity")),
                        "订单继承": _as_int(d.get("order_total_inherit")),
                        "订单共享": _as_int(d.get("order_total_share")),
                        "转化率活动": _fmt_pct(d.get("conversion_rate_activity"), 2),
                        "转化率继承": _fmt_pct(d.get("conversion_rate_inherit"), 2),
                        "转化率共享": _fmt_pct(d.get("conversion_rate_share"), 2),
                        "转化率总计": _fmt_pct(d.get("conversion_rate_total"), 2),
                    }
                )
        data = output.getvalue().encode("utf-8-sig")
        tag = "simple" if md == "simple" else "complex"
        filename = quote(f"sales_inflow_conversion_monthly_{tag}_all.csv")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出CSV失败: {e}")


# -------------------- API：投流渠道承接员工配置表 --------------------
CHANNEL_STAFF_ALLOWED_TYPES = ("活动渠道", "开户渠道")

@app.get("/api/config/channel-staff", response_model=List[ChannelStaffOut])
async def list_channel_staff() -> List[ChannelStaffOut]:
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, branch_name, staff_name, "
                f"COALESCE(NULLIF(TRIM(channel_type), ''), '开户渠道') AS channel_type, "
                f"created_at, updated_at "
                f"FROM `{CONFIG_CHANNEL_STAFF_TABLE}` ORDER BY id ASC"
            )
            rows = cur.fetchall()
        return [ChannelStaffOut(**_row_config_time_fix(row)) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/channel-staff", response_model=ChannelStaffOut)
async def create_channel_staff(body: ChannelStaffCreate, request: Request) -> ChannelStaffOut:
    _reject_if_readonly(request)
    if body.channel_type not in CHANNEL_STAFF_ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="channel_type 仅支持：活动渠道/开户渠道")
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            next_id = _next_config_id(cur, CONFIG_CHANNEL_STAFF_TABLE)
            cur.execute(
                f"INSERT INTO `{CONFIG_CHANNEL_STAFF_TABLE}` "
                f"(id, branch_name, staff_name, channel_type, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s)",
                (next_id, body.branch_name, body.staff_name, body.channel_type, now_str, now_str),
            )
            cur.execute(
                f"SELECT id, branch_name, staff_name, "
                f"COALESCE(NULLIF(TRIM(channel_type), ''), '开户渠道') AS channel_type, "
                f"created_at, updated_at "
                f"FROM `{CONFIG_CHANNEL_STAFF_TABLE}` WHERE id = %s",
                (next_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return ChannelStaffOut(**_row_config_time_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/channel-staff/{item_id}", response_model=ChannelStaffOut)
async def update_channel_staff(item_id: int, body: ChannelStaffUpdate, request: Request) -> ChannelStaffOut:
    _reject_if_readonly(request)
    if body.branch_name is None and body.staff_name is None and body.channel_type is None:
        raise HTTPException(status_code=400, detail="至少提供一个需要更新的字段")
    if body.channel_type is not None and body.channel_type not in CHANNEL_STAFF_ALLOWED_TYPES:
        raise HTTPException(status_code=400, detail="channel_type 仅支持：活动渠道/开户渠道")
    try:
        # StarRocks 部分表不支持 UPDATE（会报 1064 does not support update）
        # 这里用“查旧值 -> DELETE -> INSERT”实现修改。
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, branch_name, staff_name, "
                f"COALESCE(NULLIF(TRIM(channel_type), ''), '开户渠道') AS channel_type, "
                f"created_at, updated_at "
                f"FROM `{CONFIG_CHANNEL_STAFF_TABLE}` WHERE id = %s",
                (item_id,),
            )
            old = cur.fetchone()
            if not old:
                raise HTTPException(status_code=404, detail="记录不存在")

            new_branch = body.branch_name if body.branch_name is not None else old.get("branch_name")
            new_staff = body.staff_name if body.staff_name is not None else old.get("staff_name")
            new_channel_type = (
                body.channel_type
                if body.channel_type is not None
                else (old.get("channel_type") or "开户渠道")
            )
            created_at = _fmt_dt(old.get("created_at")) or now_str

            cur.execute(
                f"DELETE FROM `{CONFIG_CHANNEL_STAFF_TABLE}` WHERE id = %s",
                (item_id,),
            )
            cur.execute(
                f"INSERT INTO `{CONFIG_CHANNEL_STAFF_TABLE}` "
                f"(id, branch_name, staff_name, channel_type, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s)",
                (item_id, new_branch, new_staff, new_channel_type, created_at, now_str),
            )
            cur.execute(
                f"SELECT id, branch_name, staff_name, "
                f"COALESCE(NULLIF(TRIM(channel_type), ''), '开户渠道') AS channel_type, "
                f"created_at, updated_at "
                f"FROM `{CONFIG_CHANNEL_STAFF_TABLE}` WHERE id = %s",
                (item_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return ChannelStaffOut(**_row_config_time_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/channel-staff/{item_id}")
async def delete_channel_staff(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"DELETE FROM `{CONFIG_CHANNEL_STAFF_TABLE}` WHERE id = %s",
                (item_id,),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：投顾中心-销售订单配置 --------------------

def _mask_customer_phone(v: Any) -> str:
    s = str(v or "")
    if len(s) >= 11:
        return s[:3] + "****" + s[-4:]
    return s


@app.get("/api/config/sales-order")
async def list_sales_order_config(
    date: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
    month: Optional[str] = None,
) -> Dict[str, Any]:
    """
    销售订单配置列表
    - 优先按 `date=YYYY-MM-DD` 拉取“当天”数据
    - 未传 date 则优先使用 month（YYYY-MM）
    - 都没传则拉取配置表口径下“最新一日”
    """
    page = max(1, int(page or 1))
    page_size = int(page_size or 20)
    if page_size < 1:
        page_size = 20
    if page_size > 2000:
        page_size = 2000
    offset = (page - 1) * page_size

    try:
        target_day = _parse_date(date)
        # 只有用户显式传入 month 时才按月过滤；
        # month=None 不能默认成“当前月”，否则会绕过“最新一天”的兜底逻辑。
        target_month = _parse_month_yyyy_mm(month) if (not target_day and month and str(month).strip()) else None

        # 最新一日：成交日期来自配置表 pay_time（仅统计有效 pay_amount）
        if not target_day and not target_month:
            with _db_cursor() as cur:
                cur.execute(
                    f"""
                    SELECT MAX(SUBSTRING(TRIM(CAST(c.pay_time AS STRING)), 1, 10)) AS d
                    FROM `{CONFIG_SALES_ORDER_TABLE}` c
                    WHERE c.pay_time IS NOT NULL
                      AND TRIM(CAST(c.pay_time AS STRING)) != ''
                      AND c.pay_amount IS NOT NULL
                      AND CAST(c.pay_amount AS STRING) != '0.00000'
                    """
                )
                row = cur.fetchone() or {}
                raw_d = row.get("d")
                target_day = _parse_date(raw_d)
                if not target_day and raw_d:
                    s = str(raw_d).strip()
                    # 尽量截取 YYYY-MM-DD
                    target_day = s[:10] if len(s) >= 10 else _parse_date(s)

        if target_day:
            filter_sql = "SUBSTRING(TRIM(CAST(c.pay_time AS STRING)), 1, 10) = %s"
            filter_arg = target_day
            order_sql = " ORDER BY c.pay_time DESC, c.sole_code ASC"
        elif target_month:
            # 月度兜底：以 pay_time 的月份筛选（列表页）
            filter_sql = "SUBSTRING(TRIM(CAST(c.pay_time AS STRING)), 1, 7) = %s"
            filter_arg = target_month
            order_sql = " ORDER BY c.pay_time DESC, c.sole_code ASC"
        else:
            # 理论上不会走到这里：前面已尝试“最新一日”兜底。
            # 再兜底一层：按当前月返回，避免给前端空响应。
            target_month = _parse_month_yyyy_mm(None)
            filter_sql = "SUBSTRING(TRIM(CAST(c.pay_time AS STRING)), 1, 7) = %s"
            filter_arg = target_month
            order_sql = " ORDER BY c.pay_time DESC, c.sole_code ASC"

        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM `{CONFIG_SALES_ORDER_TABLE}` c
                WHERE c.pay_time IS NOT NULL
                  AND TRIM(CAST(c.pay_time AS STRING)) != ''
                  AND c.pay_amount IS NOT NULL
                  AND CAST(c.pay_amount AS STRING) != '0.00000'
                  AND {filter_sql}
                """,
                (filter_arg,),
            )
            total = int((cur.fetchone() or {}).get("total") or 0)

            cur.execute(
                f"""
                SELECT
                  c.sole_code,
                  c.customer_account,
                  c.product_name,
                  c.pay_time,
                  c.in_month,
                  c.channel,
                  c.wechat_nick,
                  c.sales_owner
                FROM `{CONFIG_SALES_ORDER_TABLE}` c
                WHERE c.pay_time IS NOT NULL
                  AND TRIM(CAST(c.pay_time AS STRING)) != ''
                  AND c.pay_amount IS NOT NULL
                  AND CAST(c.pay_amount AS STRING) != '0.00000'
                  AND {filter_sql}
                {order_sql}
                LIMIT %s OFFSET %s
                """,
                (filter_arg, page_size, offset),
            )
            rows = cur.fetchall() or []

        items = []
        for r in rows:
            d = dict(r)
            d["pay_time"] = _fmt_dt(d.get("pay_time"))
            items.append(d)

        return {"date": target_day or "", "month": target_month or "", "total": total, "items": items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/sales-order/latest-date")
async def sales_order_latest_date() -> Dict[str, str]:
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT MAX(SUBSTRING(TRIM(CAST(c.pay_time AS STRING)), 1, 10)) AS d
                FROM `{CONFIG_SALES_ORDER_TABLE}` c
                WHERE c.pay_time IS NOT NULL
                  AND TRIM(CAST(c.pay_time AS STRING)) != ''
                  AND c.pay_amount IS NOT NULL
                  AND CAST(c.pay_amount AS STRING) != '0.00000'
                """
            )
            row = cur.fetchone() or {}
        d = row.get("d")
        dd = _parse_date(d)
        if not dd and d:
            s = str(d).strip()
            dd = s[:10] if len(s) >= 10 else (dd or s)
        return {"date": dd}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/sales-order/{sole_code}/{customer_account}/{product_name}")
async def update_sales_order_config(
    sole_code: str,
    customer_account: str,
    product_name: str,
    body: Dict[str, Any],
    request: Request,
) -> Dict[str, Any]:
    _reject_if_readonly(request)
    EMPTY_CUSTOMER_ACCOUNT_TOKEN = "__EMPTY_CUSTOMER_ACCOUNT__"
    sole = (sole_code or "").strip()
    acct = (customer_account or "").strip()
    prod = (product_name or "").strip()
    # customer_account 允许为空：兼容表里 NULL/空字符串的情况
    if not sole or not prod:
        raise HTTPException(status_code=400, detail="主键字段不能为空")
    if acct == EMPTY_CUSTOMER_ACCOUNT_TOKEN:
        acct = ""
    in_month = (str(body.get("in_month") or "").strip() or None)
    channel = (str(body.get("channel") or "").strip() or None)
    wechat_nick = (str(body.get("wechat_nick") or "").strip() or None)
    sales_owner = (str(body.get("sales_owner") or "").strip() or None)
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"""
                UPDATE `{CONFIG_SALES_ORDER_TABLE}`
                SET in_month = %s,
                    channel = %s,
                    wechat_nick = %s,
                    sales_owner = %s,
                    updated_time = NOW()
                WHERE sole_code = %s
                  AND COALESCE(TRIM(customer_account), '') = %s
                  AND product_name = %s
                """,
                (in_month, channel, wechat_nick, sales_owner, sole, acct, prod),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/sales-order/detail")
async def sales_order_detail(date: Optional[str] = None, month: Optional[str] = None) -> List[Dict[str, Any]]:
    target_day = _parse_date(date)
    target_month = _parse_month_yyyy_mm(month) if not target_day else None

    # 都不传：拉最新一日
    if not target_day and not target_month:
        with _db_cursor() as cur:
            cur.execute(
                """
                SELECT MAX(SUBSTRING(TRIM(CAST(pay_time AS STRING)), 1, 10)) AS d
                FROM `mv_branch_customer_detail_final`
                WHERE pay_amount IS NOT NULL
                  AND CAST(pay_amount AS STRING) != '0.00000'
                  AND pay_time IS NOT NULL
                  AND TRIM(CAST(pay_time AS STRING)) != ''
                """
            )
            row = cur.fetchone() or {}
            target_day = _parse_date(row.get("d"))

    try:
        with _db_cursor() as cur:
            if target_day:
                day_expr = "SUBSTRING(TRIM(CAST(t.pay_time AS STRING)), 1, 10) = %s"
                cur.execute(
                    f"""
                    SELECT
                      t.customer_name,
                      t.customer_account,
                      t.customer_phone,
                      t.sole_code,
                      t.product_name,
                      t.product_type AS product_type,
                      t.product_category AS product_class,
                      t.sign_type,
                      CASE
                        WHEN t.sign_type LIKE '%%升佣%%' THEN '升佣'
                        WHEN t.sign_type LIKE '%%现金%%' THEN '现金'
                        ELSE ''
                      END AS sign_method,
                      CASE
                        WHEN t.sign_type LIKE '%%新增%%' OR t.sign_type LIKE '%%新签%%' THEN '新签'
                        WHEN t.sign_type LIKE '%%续费%%' OR t.sign_type LIKE '%%续约%%' OR t.sign_type LIKE '%%续期%%' THEN '续期'
                        WHEN t.sign_type LIKE '%%复购%%' THEN '复购'
                        ELSE ''
                      END AS sign_attr,
                      t.pay_amount,
                      t.pay_commission,
                      t.refund_amount,
                      t.curr_total_asset,
                      t.pay_time,
                      t.pay_time_end,
                      t.customer_layer,
                      t.in_month,
                      t.channel,
                      c.sales_owner AS sales_owner,
                      c.wechat_nick AS wechat_nick
                    FROM `mv_branch_customer_detail_final` t
                    LEFT JOIN `{CONFIG_SALES_ORDER_TABLE}` c
                      ON TRIM(CAST(t.sole_code AS STRING)) = TRIM(CAST(c.sole_code AS STRING))
                     AND TRIM(CAST(t.customer_account AS STRING)) = TRIM(CAST(c.customer_account AS STRING))
                     AND TRIM(CAST(t.product_name AS STRING)) = TRIM(CAST(c.product_name AS STRING))
                    WHERE t.pay_amount IS NOT NULL
                      AND CAST(t.pay_amount AS STRING) != '0.00000'
                      AND t.pay_time IS NOT NULL
                      AND TRIM(CAST(t.pay_time AS STRING)) != ''
                      AND {day_expr}
                    ORDER BY t.pay_time DESC, t.sole_code ASC
                    """,
                    (target_day,),
                )
            else:
                # 月度口径：按 pay_time 的月份筛选（成交时间 = pay_time）
                ym_expr = "SUBSTRING(TRIM(CAST(t.pay_time AS STRING)), 1, 7)"
                cur.execute(
                    f"""
                    SELECT
                      t.customer_name,
                      t.customer_account,
                      t.customer_phone,
                      t.sole_code,
                      t.product_name,
                      t.product_type AS product_type,
                      t.product_category AS product_class,
                      t.sign_type,
                      CASE
                        WHEN t.sign_type LIKE '%%升佣%%' THEN '升佣'
                        WHEN t.sign_type LIKE '%%现金%%' THEN '现金'
                        ELSE ''
                      END AS sign_method,
                      CASE
                        WHEN t.sign_type LIKE '%%新增%%' OR t.sign_type LIKE '%%新签%%' THEN '新签'
                        WHEN t.sign_type LIKE '%%续费%%' OR t.sign_type LIKE '%%续约%%' OR t.sign_type LIKE '%%续期%%' THEN '续期'
                        WHEN t.sign_type LIKE '%%复购%%' THEN '复购'
                        ELSE ''
                      END AS sign_attr,
                      t.pay_amount,
                      t.pay_commission,
                      t.refund_amount,
                      t.curr_total_asset,
                      t.pay_time,
                      t.pay_time_end,
                      t.customer_layer,
                      t.in_month,
                      t.channel,
                      c.sales_owner AS sales_owner,
                      c.wechat_nick AS wechat_nick
                    FROM `mv_branch_customer_detail_final` t
                    LEFT JOIN `{CONFIG_SALES_ORDER_TABLE}` c
                      ON TRIM(CAST(t.sole_code AS STRING)) = TRIM(CAST(c.sole_code AS STRING))
                     AND TRIM(CAST(t.customer_account AS STRING)) = TRIM(CAST(c.customer_account AS STRING))
                     AND TRIM(CAST(t.product_name AS STRING)) = TRIM(CAST(c.product_name AS STRING))
                    WHERE t.pay_amount IS NOT NULL
                      AND CAST(t.pay_amount AS STRING) != '0.00000'
                      AND t.pay_time IS NOT NULL
                      AND TRIM(CAST(t.pay_time AS STRING)) != ''
                      AND {ym_expr} = %s
                    ORDER BY t.pay_time DESC, t.sole_code ASC
                    """,
                    (target_month,),
                )
            rows = cur.fetchall() or []
        items = []
        for r in rows:
            d = dict(r)
            d["customer_phone"] = _mask_customer_phone(d.get("customer_phone"))
            d["pay_time"] = _fmt_dt(d.get("pay_time"))
            d["pay_time_end"] = _fmt_dt(d.get("pay_time_end"))
            for k in ("pay_amount", "pay_commission"):
                if d.get(k) is not None:
                    try:
                        d[k] = float(d[k])
                    except Exception:
                        pass
            for k in ("refund_amount", "curr_total_asset"):
                if d.get(k) is not None:
                    try:
                        d[k] = float(d[k])
                    except Exception:
                        pass
            # 前端“支付金额”：升佣用佣金，现金用支付金额
            try:
                sign_method = str(d.get("sign_method") or "")
                if sign_method == "升佣":
                    d["pay_amount_display"] = d.get("pay_commission")
                else:
                    d["pay_amount_display"] = d.get("pay_amount")
            except Exception:
                d["pay_amount_display"] = None
            items.append(d)
        return items
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/sales-order/summary")
async def sales_order_summary(date: Optional[str] = None, month: Optional[str] = None) -> List[Dict[str, Any]]:
    target_day = _parse_date(date)
    target_month = _parse_month_yyyy_mm(month) if not target_day else None

    if not target_day and not target_month:
        with _db_cursor() as cur:
            cur.execute(
                """
                SELECT MAX(SUBSTRING(TRIM(CAST(pay_time AS STRING)), 1, 10)) AS d
                FROM `mv_branch_customer_detail_final`
                WHERE pay_amount IS NOT NULL
                  AND CAST(pay_amount AS STRING) != '0.00000'
                  AND pay_time IS NOT NULL
                  AND TRIM(CAST(pay_time AS STRING)) != ''
                """
            )
            row = cur.fetchone() or {}
            target_day = _parse_date(row.get("d"))

    try:
        with _db_cursor() as cur:
            if target_day:
                day_expr = "SUBSTRING(TRIM(CAST(t.pay_time AS STRING)), 1, 10) = %s"
                cur.execute(
                    f"""
                    SELECT
                      COALESCE(NULLIF(TRIM(CAST(c.sales_owner AS STRING)), ''), '-') AS sales_owner,
                      COUNT(*) AS total_count,
                      SUM(CASE WHEN t.sign_type LIKE '%%升佣%%' THEN 1 ELSE 0 END) AS commission_count,
                      SUM(CASE WHEN t.sign_type LIKE '%%现金%%' THEN 1 ELSE 0 END) AS cash_count,
                      SUM(CASE WHEN t.sign_type LIKE '%%现金%%' THEN t.pay_amount ELSE 0 END) AS cash_amount,
                      -- 新签/续期/复购：按 customer_layer 统计（mv 口径）
                      SUM(CASE WHEN t.customer_layer = '新增' THEN 1 ELSE 0 END) AS new_count,
                      SUM(CASE WHEN t.customer_layer = '续费' THEN 1 ELSE 0 END) AS renew_count,
                      SUM(CASE WHEN t.customer_layer = '复购' THEN 1 ELSE 0 END) AS repurchase_count
                    FROM `mv_branch_customer_detail_final` t
                    LEFT JOIN `{CONFIG_SALES_ORDER_TABLE}` c
                      ON TRIM(CAST(t.sole_code AS STRING)) = TRIM(CAST(c.sole_code AS STRING))
                     AND TRIM(CAST(t.customer_account AS STRING)) = TRIM(CAST(c.customer_account AS STRING))
                     AND TRIM(CAST(t.product_name AS STRING)) = TRIM(CAST(c.product_name AS STRING))
                    WHERE t.pay_amount IS NOT NULL
                      AND CAST(t.pay_amount AS STRING) != '0.00000'
                      AND t.pay_time IS NOT NULL
                      AND TRIM(CAST(t.pay_time AS STRING)) != ''
                      AND {day_expr}
                    GROUP BY COALESCE(NULLIF(TRIM(CAST(c.sales_owner AS STRING)), ''), '-')
                    ORDER BY total_count DESC, sales_owner ASC
                    """,
                    (target_day,),
                )
            else:
                # 月度口径：按 pay_time 的月份筛选（成交时间 = pay_time）
                ym_expr = "SUBSTRING(TRIM(CAST(t.pay_time AS STRING)), 1, 7)"
                cur.execute(
                    f"""
                    SELECT
                      COALESCE(NULLIF(TRIM(CAST(c.sales_owner AS STRING)), ''), '-') AS sales_owner,
                      COUNT(*) AS total_count,
                      SUM(CASE WHEN t.sign_type LIKE '%%升佣%%' THEN 1 ELSE 0 END) AS commission_count,
                      SUM(CASE WHEN t.sign_type LIKE '%%现金%%' THEN 1 ELSE 0 END) AS cash_count,
                      SUM(CASE WHEN t.sign_type LIKE '%%现金%%' THEN t.pay_amount ELSE 0 END) AS cash_amount,
                      -- 新签/续期/复购：按 customer_layer 统计（mv 口径）
                      SUM(CASE WHEN t.customer_layer = '新增' THEN 1 ELSE 0 END) AS new_count,
                      SUM(CASE WHEN t.customer_layer = '续费' THEN 1 ELSE 0 END) AS renew_count,
                      SUM(CASE WHEN t.customer_layer = '复购' THEN 1 ELSE 0 END) AS repurchase_count
                    FROM `mv_branch_customer_detail_final` t
                    LEFT JOIN `{CONFIG_SALES_ORDER_TABLE}` c
                      ON TRIM(CAST(t.sole_code AS STRING)) = TRIM(CAST(c.sole_code AS STRING))
                     AND TRIM(CAST(t.customer_account AS STRING)) = TRIM(CAST(c.customer_account AS STRING))
                     AND TRIM(CAST(t.product_name AS STRING)) = TRIM(CAST(c.product_name AS STRING))
                    WHERE t.pay_amount IS NOT NULL
                      AND CAST(t.pay_amount AS STRING) != '0.00000'
                      AND t.pay_time IS NOT NULL
                      AND TRIM(CAST(t.pay_time AS STRING)) != ''
                      AND {ym_expr} = %s
                    GROUP BY COALESCE(NULLIF(TRIM(CAST(c.sales_owner AS STRING)), ''), '-')
                    ORDER BY total_count DESC, sales_owner ASC
                    """,
                    (target_month,),
                )
            rows = cur.fetchall() or []

        items: List[Dict[str, Any]] = []
        totals = {
            "commission_count": 0,
            "cash_count": 0,
            "total_count": 0,
            "cash_amount": 0.0,
            "new_count": 0,
            "renew_count": 0,
            "repurchase_count": 0,
        }
        for r in rows:
            d = dict(r)
            for k in ("commission_count", "cash_count", "total_count", "new_count", "renew_count", "repurchase_count"):
                try:
                    d[k] = int(d.get(k) or 0)
                except Exception:
                    d[k] = 0
            try:
                d["cash_amount"] = float(d.get("cash_amount") or 0)
            except Exception:
                d["cash_amount"] = 0.0
            items.append(d)
            totals["commission_count"] += d["commission_count"]
            totals["cash_count"] += d["cash_count"]
            totals["total_count"] += d["total_count"]
            totals["cash_amount"] += d["cash_amount"]
            totals["new_count"] += d["new_count"]
            totals["renew_count"] += d["renew_count"]
            totals["repurchase_count"] += d["repurchase_count"]

        items.append(
            {
                "sales_owner": "总计",
                "commission_count": totals["commission_count"],
                "cash_count": totals["cash_count"],
                "total_count": totals["total_count"],
                "cash_amount": totals["cash_amount"],
                "new_count": totals["new_count"],
                "renew_count": totals["renew_count"],
                "repurchase_count": totals["repurchase_count"],
            }
        )
        return items
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/sales-order/detail/export.csv")
async def sales_order_detail_export_csv(date: Optional[str] = None, month: Optional[str] = None):
    try:
        # 导出不复用 detail 接口，避免时间格式化/金额改口径；直接输出原始字段
        target_day = _parse_date(date)
        target_month = _parse_month_yyyy_mm(month) if not target_day else None

        if not target_day and not target_month:
            with _db_cursor() as cur:
                cur.execute(
                    """
                    SELECT MAX(SUBSTRING(TRIM(CAST(pay_time AS STRING)), 1, 10)) AS d
                    FROM `mv_branch_customer_detail_final`
                    WHERE pay_amount IS NOT NULL
                      AND CAST(pay_amount AS STRING) != '0.00000'
                      AND pay_time IS NOT NULL
                      AND TRIM(CAST(pay_time AS STRING)) != ''
                    """
                )
                row = cur.fetchone() or {}
                target_day = _parse_date(row.get("d"))

        with _db_cursor() as cur:
            if target_day:
                day_expr = "SUBSTRING(TRIM(CAST(t.pay_time AS STRING)), 1, 10) = %s"
                cur.execute(
                    f"""
                    SELECT
                      t.customer_name,
                      t.customer_account,
                      t.customer_phone,
                      t.sole_code,
                      t.product_name,
                      t.product_type AS product_type,
                      t.product_category AS product_class,
                      t.sign_type,
                      t.pay_amount,
                      t.pay_commission,
                      t.refund_amount,
                      t.curr_total_asset,
                      t.pay_time,
                      t.pay_time_end,
                      t.customer_layer,
                      t.in_month,
                      t.channel,
                      c.wechat_nick AS wechat_nick,
                      c.sales_owner AS sales_owner
                    FROM `mv_branch_customer_detail_final` t
                    LEFT JOIN `{CONFIG_SALES_ORDER_TABLE}` c
                      ON t.sole_code = c.sole_code
                     AND t.customer_account = c.customer_account
                     AND t.product_name = c.product_name
                    WHERE t.pay_amount IS NOT NULL
                      AND CAST(t.pay_amount AS STRING) != '0.00000'
                      AND t.pay_time IS NOT NULL
                      AND TRIM(CAST(t.pay_time AS STRING)) != ''
                      AND {day_expr}
                    ORDER BY t.pay_time DESC, t.sole_code ASC
                    """,
                    (target_day,),
                )
            else:
                ym_expr = "SUBSTRING(TRIM(CAST(t.pay_time AS STRING)), 1, 7)"
                cur.execute(
                    f"""
                    SELECT
                      t.customer_name,
                      t.customer_account,
                      t.customer_phone,
                      t.sole_code,
                      t.product_name,
                      t.product_type AS product_type,
                      t.product_category AS product_class,
                      t.sign_type,
                      t.pay_amount,
                      t.pay_commission,
                      t.refund_amount,
                      t.curr_total_asset,
                      t.pay_time,
                      t.pay_time_end,
                      t.customer_layer,
                      t.in_month,
                      t.channel,
                      c.wechat_nick AS wechat_nick,
                      c.sales_owner AS sales_owner
                    FROM `mv_branch_customer_detail_final` t
                    LEFT JOIN `{CONFIG_SALES_ORDER_TABLE}` c
                      ON t.sole_code = c.sole_code
                     AND t.customer_account = c.customer_account
                     AND t.product_name = c.product_name
                    WHERE t.pay_amount IS NOT NULL
                      AND CAST(t.pay_amount AS STRING) != '0.00000'
                      AND t.pay_time IS NOT NULL
                      AND TRIM(CAST(t.pay_time AS STRING)) != ''
                      AND {ym_expr} = %s
                    ORDER BY t.pay_time DESC, t.sole_code ASC
                    """,
                    (target_month,),
                )
            rows = cur.fetchall() or []

        cols = [
            ("客户姓名", "customer_name"),
            ("资金账号", "customer_account"),
            ("手机号", "customer_phone"),
            ("订单编号", "sole_code"),
            ("产品名称", "product_name"),
            ("产品类型", "product_type"),
            ("产品归类", "product_class"),
            ("签约方式", "sign_type"),
            ("支付金额", "pay_amount"),
            ("佣金", "pay_commission"),
            ("退款金额", "refund_amount"),
            ("总资产", "curr_total_asset"),
            ("成交时间", "pay_time"),
            ("支付结束时间", "pay_time_end"),
            ("客户分层", "customer_layer"),
            ("进线月份", "in_month"),
            ("渠道", "channel"),
            ("微信昵称", "wechat_nick"),
            ("销售归属", "sales_owner"),
        ]

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[c[0] for c in cols])
        writer.writeheader()
        for r in rows:
            d = dict(r)
            out: Dict[str, Any] = {}
            for cn, en in cols:
                v = d.get(en)
                if en == "customer_phone":
                    v = _mask_customer_phone(v)
                if en == "in_month" and v is not None:
                    s = str(v).strip()
                    # 清理导出时可能出现的前导标点，统一输出 YYYY-MM。
                    s = s.lstrip(",，' ")
                    m = _re.search(r"\d{4}-\d{2}", s)
                    if m:
                        ym = m.group(0)
                        # 用公式文本形式导出，避免 Excel 自动显示为 Apr-26。
                        v = f'="{ym}"'
                    else:
                        v = s
                out[cn] = v
            writer.writerow(out)

        data = output.getvalue().encode("utf-8-sig")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename={quote(f'sales_order_detail_{_parse_date(date) or _parse_month_yyyy_mm(month)}.csv')}"},
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"导出明细CSV失败: {e}")


@app.get("/api/config/sales-order/summary/export.csv")
async def sales_order_summary_export_csv(date: Optional[str] = None, month: Optional[str] = None):
    rows = await sales_order_summary(date=date, month=month)
    cols = [
        "sales_owner",
        "commission_count",
        "cash_count",
        "total_count",
        "cash_amount",
        "new_count",
        "renew_count",
        "repurchase_count",
    ]
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=cols)
    writer.writeheader()
    for r in rows:
        writer.writerow({k: r.get(k) for k in cols})
    data = output.getvalue().encode("utf-8-sig")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={quote(f'sales_order_summary_{_parse_date(date) or _parse_month_yyyy_mm(month)}.csv')}"},
    )


# -------------------- API：投顾中心-签约客户群管理配置 --------------------

@app.get("/api/config/sign-customer-group")
async def list_sign_customer_group(
    month: Optional[str] = None,
    page: int = 1,
    page_size: int = 20,
) -> Dict[str, Any]:
    m = _parse_month_yyyy_mm(month)
    page = max(1, int(page or 1))
    page_size = int(page_size or 20)
    if page_size < 1:
        page_size = 20
    if page_size > 200:
        page_size = 200
    offset = (page - 1) * page_size
    ym_expr = "SUBSTRING(TRIM(CAST(pay_time_end AS STRING)), 1, 7)"

    try:
        with _db_cursor() as cur:
            cur.execute(
                f"""
                SELECT COUNT(*) AS total
                FROM `{CONFIG_SIGN_CUSTOMER_GROUP_TABLE}`
                WHERE pay_time_end IS NOT NULL
                  AND TRIM(CAST(pay_time_end AS STRING)) != ''
                  AND {ym_expr} = %s
                """,
                (m,),
            )
            total = int((cur.fetchone() or {}).get("total") or 0)

            cur.execute(
                f"""
                SELECT
                  sole_code, customer_name, customer_phone, customer_account,
                  wechat_nick, pay_time, sign_type, curr_total_asset,
                  pay_time_end, refund_amount, in_group, updated_time
                FROM `{CONFIG_SIGN_CUSTOMER_GROUP_TABLE}`
                WHERE pay_time_end IS NOT NULL
                  AND TRIM(CAST(pay_time_end AS STRING)) != ''
                  AND {ym_expr} = %s
                ORDER BY pay_time DESC, sole_code ASC
                LIMIT %s OFFSET %s
                """,
                (m, page_size, offset),
            )
            rows = cur.fetchall() or []

        items: List[Dict[str, Any]] = []
        for r in rows:
            d = dict(r)
            for k in ("pay_time", "pay_time_end", "updated_time"):
                if k in d:
                    d[k] = _fmt_dt(d.get(k))
            for k in ("curr_total_asset", "refund_amount"):
                if d.get(k) is not None:
                    try:
                        d[k] = float(d[k])
                    except Exception:
                        pass
            if d.get("in_group") is not None:
                try:
                    d["in_group"] = int(d["in_group"])
                except Exception:
                    pass
            items.append(d)

        return {"month": m, "total": total, "items": items}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/sign-customer-group/{sole_code}/{customer_account}")
async def update_sign_customer_group_in_group(
    sole_code: str,
    customer_account: str,
    body: SignCustomerGroupUpdate,
    request: Request,
) -> Dict[str, Any]:
    _reject_if_readonly(request)
    in_group = 1 if int(body.in_group or 0) == 1 else 0
    sole = (sole_code or "").strip()
    acct = (customer_account or "").strip()
    if not sole or not acct:
        raise HTTPException(status_code=400, detail="sole_code 与 customer_account 不能为空")

    try:
        with _db_cursor() as cur:
            cur.execute(
                f"""
                UPDATE `{CONFIG_SIGN_CUSTOMER_GROUP_TABLE}`
                SET in_group = %s, updated_time = NOW()
                WHERE sole_code = %s AND customer_account = %s
                """,
                (in_group, sole, acct),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：客户中心-商机线索配置 --------------------

@app.get("/api/config/opportunity-leads", response_model=List[OpportunityLeadOut])
async def list_opportunity_leads() -> List[OpportunityLeadOut]:
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, biz_category_big, biz_category_small, clue_name, is_important, remark, table_name, created_at, updated_at "
                f"FROM `{CONFIG_OPPORTUNITY_LEAD_TABLE}` ORDER BY id ASC"
            )
            rows = cur.fetchall()
        return [OpportunityLeadOut(**_row_opportunity_lead_fix(r)) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/opportunity-leads", response_model=OpportunityLeadOut)
async def create_opportunity_lead(body: OpportunityLeadCreate, request: Request) -> OpportunityLeadOut:
    _reject_if_readonly(request)
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"INSERT INTO `{CONFIG_OPPORTUNITY_LEAD_TABLE}` "
                f"(biz_category_big, biz_category_small, clue_name, is_important, remark, table_name, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                (
                    (body.biz_category_big or "").strip() or None,
                    (body.biz_category_small or "").strip() or None,
                    (body.clue_name or "").strip() or None,
                    (1 if body.is_important else 0) if body.is_important is not None else None,
                    (body.remark or "").strip() or None,
                    (body.table_name or "").strip() or None,
                    now_str,
                    now_str,
                ),
            )
            # StarRocks 对 lastrowid 支持有限，这里用倒序取最新一条作为简单方案
            cur.execute(
                f"SELECT id, biz_category_big, biz_category_small, clue_name, is_important, remark, table_name, created_at, updated_at "
                f"FROM `{CONFIG_OPPORTUNITY_LEAD_TABLE}` ORDER BY id DESC LIMIT 1"
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return OpportunityLeadOut(**_row_opportunity_lead_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/opportunity-leads/{item_id}", response_model=OpportunityLeadOut)
async def update_opportunity_lead(item_id: int, body: OpportunityLeadUpdate, request: Request) -> OpportunityLeadOut:
    _reject_if_readonly(request)
    if (
        body.biz_category_big is None
        and body.biz_category_small is None
        and body.clue_name is None
        and body.is_important is None
        and body.remark is None
        and body.table_name is None
    ):
        raise HTTPException(status_code=400, detail="至少提供一个需要更新的字段")
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, biz_category_big, biz_category_small, clue_name, is_important, remark, table_name, created_at, updated_at "
                f"FROM `{CONFIG_OPPORTUNITY_LEAD_TABLE}` WHERE id = %s",
                (item_id,),
            )
            old = cur.fetchone()
            if not old:
                raise HTTPException(status_code=404, detail="记录不存在")

            def _pick(key: str, new_val: Any):
                return new_val if new_val is not None else old.get(key)

            created_at = _fmt_dt(old.get("created_at")) or now_str
            new_big = (str(_pick("biz_category_big", body.biz_category_big) or "").strip() or None)
            new_small = (str(_pick("biz_category_small", body.biz_category_small) or "").strip() or None)
            new_clue = (str(_pick("clue_name", body.clue_name) or "").strip() or None)
            new_imp = _pick(
                "is_important",
                (1 if body.is_important else 0) if body.is_important is not None else None,
            )
            new_remark = (str(_pick("remark", body.remark) or "").strip() or None)
            new_table = (str(_pick("table_name", body.table_name) or "").strip() or None)

            # StarRocks 部分表不支持 UPDATE：用 DELETE + INSERT
            cur.execute(f"DELETE FROM `{CONFIG_OPPORTUNITY_LEAD_TABLE}` WHERE id = %s", (item_id,))
            cur.execute(
                f"INSERT INTO `{CONFIG_OPPORTUNITY_LEAD_TABLE}` "
                f"(id, biz_category_big, biz_category_small, clue_name, is_important, remark, table_name, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (item_id, new_big, new_small, new_clue, new_imp, new_remark, new_table, created_at, now_str),
            )
            cur.execute(
                f"SELECT id, biz_category_big, biz_category_small, clue_name, is_important, remark, table_name, created_at, updated_at "
                f"FROM `{CONFIG_OPPORTUNITY_LEAD_TABLE}` WHERE id = %s",
                (item_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return OpportunityLeadOut(**_row_opportunity_lead_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/opportunity-leads/{item_id}")
async def delete_opportunity_lead(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(f"DELETE FROM `{CONFIG_OPPORTUNITY_LEAD_TABLE}` WHERE id = %s", (item_id,))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：客户中心-早盘人气股战绩追踪配置 --------------------

@app.get("/api/config/morning-hot-stock-track/tg-names", response_model=List[str])
async def list_morning_hot_stock_track_tg_names() -> List[str]:
    """
    返回老师列表（去重），用于前端下拉筛选。
    """
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT DISTINCT tg_name AS tg_name "
                f"FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` "
                f"WHERE tg_name IS NOT NULL AND TRIM(tg_name) != '' "
                f"ORDER BY tg_name"
            )
            rows = cur.fetchall() or []
        return [str(r.get("tg_name") or "").strip() for r in rows if str(r.get("tg_name") or "").strip()]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/api/config/morning-hot-stock-track", response_model=List[MorningHotStockTrackOut])
async def list_morning_hot_stock_track(tg_name: Optional[str] = None) -> List[MorningHotStockTrackOut]:
    try:
        with _db_cursor() as cur:
            if tg_name and tg_name.strip():
                cur.execute(
                    f"SELECT id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at "
                    f"FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` "
                    f"WHERE tg_name = %s "
                    f"ORDER BY created_at DESC, id DESC",
                    (tg_name.strip(),),
                )
            else:
                cur.execute(
                    f"SELECT id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at "
                    f"FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` "
                    f"ORDER BY created_at DESC, id DESC"
                )
            rows = cur.fetchall()
        return [MorningHotStockTrackOut(**_row_morning_hot_stock_track_fix(r)) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/morning-hot-stock-track/performance")
async def morning_hot_stock_track_performance(
    tg_name: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> Dict[str, Any]:
    """
    客户中心-早盘人气股：战绩统计（来自 mv_morning_hot_stock_track）

    - 默认日期范围：上个月（自然月）
    - 胜率口径：
      - 次日涨跌幅胜率：highprice_next_date > openprice 记为涨（1），否则 0；空值不计胜
      - T+3 日内最高涨跌幅胜率：highprice_third_next_date > openprice 记为涨（1），否则 0；空值不计胜
    """
    name = (tg_name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="tg_name 必填")

    # 默认上个月自然月
    if not start_date or not start_date.strip() or not end_date or not end_date.strip():
        today = date.today()
        first_this_month = today.replace(day=1)
        last_prev_month = first_this_month - timedelta(days=1)
        first_prev_month = last_prev_month.replace(day=1)
        start_date = first_prev_month.strftime("%Y-%m-%d")
        end_date = last_prev_month.strftime("%Y-%m-%d")
    else:
        start_date = start_date.strip()
        end_date = end_date.strip()

    try:
        with _db_cursor() as cur:
            sql = (
                "SELECT "
                "  biz_date, stock_name, stock_code, openprice, "
                "  highprice_next_date, highprice_third_next_date, "
                "  next_day_high_chg_pct, t3_high_chg_pct, "
                "  CASE "
                "    WHEN openprice IS NULL OR openprice = 0 OR highprice_next_date IS NULL THEN NULL "
                "    WHEN highprice_next_date > openprice THEN 1 ELSE 0 "
                "  END AS next_win, "
                "  CASE "
                "    WHEN openprice IS NULL OR openprice = 0 OR highprice_third_next_date IS NULL THEN NULL "
                "    WHEN highprice_third_next_date > openprice THEN 1 ELSE 0 "
                "  END AS t3_win "
                "FROM mv_morning_hot_stock_track "
                "WHERE tg_name = %s AND biz_date BETWEEN %s AND %s "
                "ORDER BY biz_date ASC, stock_code ASC"
            )
            cur.execute(sql, (name, start_date, end_date))
            rows = cur.fetchall() or []

        total = len(rows)
        next_wins = sum(1 for r in rows if (r.get("next_win") == 1 or str(r.get("next_win") or "") == "1"))
        t3_wins = sum(1 for r in rows if (r.get("t3_win") == 1 or str(r.get("t3_win") or "") == "1"))

        def _pct(w: int, n: int) -> Optional[str]:
            if n <= 0:
                return None
            return f"{round((w / n) * 100, 2)}%"

        # 明细：日期统一为 YYYY-MM-DD（避免前端 GMT 展示）
        items = []
        for r in rows:
            d = _parse_date(r.get("biz_date"))
            items.append(
                {
                    "biz_date": d,
                    "stock_name": r.get("stock_name"),
                    "stock_code": r.get("stock_code"),
                    "openprice": r.get("openprice"),
                    "t1_pct": r.get("next_day_high_chg_pct"),
                    "t3_pct": r.get("t3_high_chg_pct"),
                }
            )

        return {
            "tg_name": name,
            "start_date": start_date,
            "end_date": end_date,
            "push_count": total,
            "next_win_rate": _pct(next_wins, total),
            "t3_win_rate": _pct(t3_wins, total),
            "items": items,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/morning-hot-stock-track/performance/export.csv")
async def morning_hot_stock_track_performance_export_csv(
    tg_name: str,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
):
    """
    导出“战绩”明细（CSV）。
    字段与前端明细表一致：日期（买入）、人气股、开盘价、T+1涨幅、T+3最高涨幅。
    """
    data = await morning_hot_stock_track_performance(tg_name=tg_name, start_date=start_date, end_date=end_date)
    items = data.get("items") or []

    sio = io.StringIO()
    w = csv.writer(sio)
    w.writerow(["日期（买入）", "人气股", "开盘价", "T+1涨幅", "T+3最高涨幅"])
    for r in items:
        w.writerow(
            [
                r.get("biz_date") or "",
                r.get("stock_name") or "",
                r.get("openprice") if r.get("openprice") is not None else "",
                r.get("t1_pct") or "",
                r.get("t3_pct") or "",
            ]
        )

    # Excel 兼容：UTF-8 with BOM
    payload = ("\ufeff" + sio.getvalue()).encode("utf-8")
    fname = f"{data.get('tg_name')}_早评人气股_{data.get('start_date')}~{data.get('end_date')}.csv"
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"}
    return StreamingResponse(io.BytesIO(payload), media_type="text/csv; charset=utf-8", headers=headers)


@app.post("/api/config/morning-hot-stock-track", response_model=MorningHotStockTrackOut)
async def create_morning_hot_stock_track(body: MorningHotStockTrackCreate, request: Request) -> MorningHotStockTrackOut:
    _reject_if_readonly(request)
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"INSERT INTO `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` "
                f"(id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at) "
                f"VALUES (DEFAULT, %s, %s, %s, %s, %s, %s, %s)",
                (
                    (body.tg_name or "").strip() or None,
                    (body.biz_date or "").strip() or None,
                    (body.stock_name or "").strip() or None,
                    (body.stock_code or "").strip() or None,
                    (body.remark or "").strip() or None,
                    now_str,
                    now_str,
                ),
            )
            # 取最新一条（倒序）
            cur.execute(
                f"SELECT id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at "
                f"FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` ORDER BY id DESC LIMIT 1"
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return MorningHotStockTrackOut(**_row_morning_hot_stock_track_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/morning-hot-stock-track/{item_id}", response_model=MorningHotStockTrackOut)
async def update_morning_hot_stock_track(item_id: int, body: MorningHotStockTrackUpdate, request: Request) -> MorningHotStockTrackOut:
    _reject_if_readonly(request)
    if body.tg_name is None and body.biz_date is None and body.stock_name is None and body.stock_code is None and body.remark is None:
        raise HTTPException(status_code=400, detail="至少提供一个需要更新的字段")
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at "
                f"FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` WHERE id = %s",
                (item_id,),
            )
            old = cur.fetchone()
            if not old:
                raise HTTPException(status_code=404, detail="记录不存在")

            def _pick(key: str, new_val: Any):
                return new_val if new_val is not None else old.get(key)

            created_at = _fmt_dt(old.get("created_at")) or now_str
            new_tg_name = (str(_pick("tg_name", body.tg_name) or "").strip() or None)
            new_biz_date = (str(_pick("biz_date", body.biz_date) or "").strip() or None)
            new_name = (str(_pick("stock_name", body.stock_name) or "").strip() or None)
            new_code = (str(_pick("stock_code", body.stock_code) or "").strip() or None)
            new_remark = (str(_pick("remark", body.remark) or "").strip() or None)

            # StarRocks 兼容：DELETE + INSERT
            cur.execute(f"DELETE FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` WHERE id = %s", (item_id,))
            cur.execute(
                f"INSERT INTO `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` "
                f"(id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s)",
                (item_id, new_tg_name, new_biz_date, new_name, new_code, new_remark, created_at, now_str),
            )
            cur.execute(
                f"SELECT id, tg_name, biz_date, stock_name, stock_code, remark, created_at, updated_at "
                f"FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` WHERE id = %s",
                (item_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return MorningHotStockTrackOut(**_row_morning_hot_stock_track_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/morning-hot-stock-track/{item_id}")
async def delete_morning_hot_stock_track(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(f"DELETE FROM `{CONFIG_MORNING_HOT_STOCK_TRACK_TABLE}` WHERE id = %s", (item_id,))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：渠道映射/消耗（code_mapping） --------------------

@app.get("/api/config/code-mapping", response_model=List[CodeMappingOut])
async def list_code_mapping() -> List[CodeMappingOut]:
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, code_value, description, stat_cost, channel_name, created_time "
                f"FROM `{CONFIG_CODE_MAPPING_TABLE}` ORDER BY id ASC"
            )
            rows = cur.fetchall()
        return [CodeMappingOut(**_row_code_mapping_fix(row)) for row in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/code-mapping", response_model=CodeMappingOut)
async def create_code_mapping(body: CodeMappingCreate, request: Request) -> CodeMappingOut:
    _reject_if_readonly(request)
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(f"SELECT 1 FROM `{CONFIG_CODE_MAPPING_TABLE}` WHERE id = %s LIMIT 1", (body.id,))
            if cur.fetchone():
                raise HTTPException(status_code=400, detail="ID 已存在")
            cur.execute(
                f"INSERT INTO `{CONFIG_CODE_MAPPING_TABLE}` "
                f"(id, code_value, description, stat_cost, channel_name, created_time) "
                f"VALUES (%s, %s, %s, %s, %s, %s)",
                (
                    int(body.id),
                    body.code_value,
                    body.description,
                    body.stat_cost,
                    body.channel_name,
                    now_str,
                ),
            )
            cur.execute(
                f"SELECT id, code_value, description, stat_cost, channel_name, created_time "
                f"FROM `{CONFIG_CODE_MAPPING_TABLE}` WHERE id = %s",
                (int(body.id),),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return CodeMappingOut(**_row_code_mapping_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/code-mapping/{item_id}", response_model=CodeMappingOut)
async def update_code_mapping(item_id: int, body: CodeMappingUpdate, request: Request) -> CodeMappingOut:
    _reject_if_readonly(request)
    if body.code_value is None and body.description is None and body.stat_cost is None and body.channel_name is None:
        raise HTTPException(status_code=400, detail="至少提供一个需要更新的字段")
    try:
        fields = []
        params: List[Any] = []
        if body.code_value is not None:
            fields.append("code_value = %s")
            params.append(body.code_value)
        if body.description is not None:
            fields.append("description = %s")
            params.append(body.description)
        if body.stat_cost is not None:
            fields.append("stat_cost = %s")
            params.append(body.stat_cost)
        if body.channel_name is not None:
            fields.append("channel_name = %s")
            params.append(body.channel_name)

        with _db_cursor() as cur:
            sql = f"UPDATE `{CONFIG_CODE_MAPPING_TABLE}` SET {', '.join(fields)} WHERE id = %s"
            params.append(int(item_id))
            cur.execute(sql, tuple(params))
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
            cur.execute(
                f"SELECT id, code_value, description, stat_cost, channel_name, created_time "
                f"FROM `{CONFIG_CODE_MAPPING_TABLE}` WHERE id = %s",
                (int(item_id),),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return CodeMappingOut(**_row_code_mapping_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/code-mapping/{item_id}")
async def delete_code_mapping(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"DELETE FROM `{CONFIG_CODE_MAPPING_TABLE}` WHERE id = %s",
                (int(item_id),),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：股票仓位/买卖配置表 --------------------

@app.get("/api/config/stock-position/products", response_model=List[str])
async def list_stock_position_products() -> List[str]:
    """返回所有不重复的产品名称，用于筛选下拉。"""
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT DISTINCT product_name FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE product_name IS NOT NULL AND product_name != '' "
                f"ORDER BY product_name"
            )
            raw_names = [r["product_name"] for r in cur.fetchall()]

            # 统一清洗 + 去重（确保顺序可控）
            cleaned: List[str] = []
            seen = set()
            for n in raw_names:
                s = (n or "").strip()
                if not s:
                    continue
                if s in seen:
                    continue
                seen.add(s)
                cleaned.append(s)

            # 过滤测试类产品名，避免污染“投顾中心-产品净值”筛选下拉
            cleaned = [n for n in cleaned if not _is_test_product_name(n)]

            # 固定下拉顺序：常用产品置顶，其余按名称稳定排序（大小写不敏感）
            pinned = {"短线王"}

            def _sort_key(x: str):
                return (0 if x in pinned else 1, x.casefold())

            return sorted(cleaned, key=_sort_key)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/stock-position", response_model=Dict[str, Any])
async def list_stock_position(
    product_names: Optional[str] = None,
    page: int = 1,
    page_size: int = 30,
) -> Dict[str, Any]:
    """
    按产品名称筛选，按日期倒序，分页。
    product_names 为单个产品名称；不传时默认展示产品名为「短线王」的数据。
    """
    if page < 1:
        page = 1
    if page_size < 1 or page_size > 500:
        page_size = 30
    if product_names is None or not product_names.strip():
        names = ["短线王"]
    else:
        first = product_names.split(",")[0].strip()
        names = [first] if first else ["短线王"]
    if not names:
        return {"total": 0, "items": [], "page": page, "page_size": page_size}

    try:
        with _db_cursor() as cur:
            placeholders = ", ".join(["%s"] * len(names))
            count_sql = (
                f"SELECT COUNT(*) AS cnt FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE TRIM(product_name) IN ({placeholders})"
            )
            cur.execute(count_sql, tuple(names))
            total = (cur.fetchone() or {}).get("cnt") or 0

            offset = (page - 1) * page_size
            list_sql = (
                f"SELECT id, product_name, trade_date, row_id, stock_code, stock_name, "
                f"position_pct, side, price, created_at, updated_at "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE TRIM(product_name) IN ({placeholders}) "
                f"ORDER BY trade_date desc, created_at DESC, id DESC LIMIT %s OFFSET %s"
            )
            cur.execute(list_sql, tuple(names) + (page_size, offset))
            rows = cur.fetchall()

        items = [StockPositionOut(**_row_stock_position_fix(r)) for r in rows]
        return {"total": total, "items": items, "page": page, "page_size": page_size}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/stock-position/export", response_model=List[StockPositionOut])
async def export_stock_position(
    product_name: Optional[str] = None,
) -> List[StockPositionOut]:
    """
    导出指定产品名称的全部股票仓位/买卖记录（不分页）。
    """
    if product_name is None or not product_name.strip():
        first = "短线王"
    else:
        first = product_name.split(",")[0].strip() or "短线王"

    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, product_name, trade_date, row_id, stock_code, stock_name, "
                f"position_pct, side, price, created_at, updated_at "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE product_name = %s "
                f"ORDER BY created_at DESC, id DESC",
                (first,),
            )
            rows = cur.fetchall()
        return [StockPositionOut(**_row_stock_position_fix(r)) for r in rows]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/stock-position/export.xlsx")
async def export_stock_position_excel(
    product_name: Optional[str] = None,
):
    """
    下载 Excel（两个 sheet）：
    - Sheet1: 当前筛选产品下的仓位明细（全量，不分页）
    - Sheet2: 该产品的净值序列（product_nav_daily_detail row_type=3）
    """
    name = (product_name or "").strip() or "短线王"

    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, product_name, trade_date, row_id, stock_code, stock_name, "
                f"position_pct, side, price, created_at, updated_at "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE product_name = %s "
                f"ORDER BY created_at DESC, id DESC",
                (name,),
            )
            rows1 = cur.fetchall()

            cur.execute(
                f"SELECT biz_date, nav, hs300_nav "
                f"FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                f"WHERE product_name = %s AND row_type = 3 "
                f"ORDER BY biz_date ASC",
                (name,),
            )
            rows2 = cur.fetchall()

            # Sheet3: 净值明细（row_type=2）
            cur.execute(
                f"SELECT biz_date, stock_name, stock_code, position_after "
                f"FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                f"WHERE product_name = %s AND row_type = 2 "
                f"  AND biz_date IS NOT NULL "
                f"  AND TRIM(CAST(biz_date AS STRING)) != '' "
                f"ORDER BY biz_date DESC, stock_code ASC",
                (name,),
            )
            rows3 = cur.fetchall()

        from openpyxl import Workbook

        wb = Workbook()

        # Sheet1: 当前页仓位明细
        ws1 = wb.active
        ws1.title = "仓位明细"
        ws1.append(["ID", "产品名称", "日期", "序号", "股票代码", "个股", "仓位(%)", "买入/卖出", "成交价", "创建时间", "更新时间"])
        for r in rows1:
            ws1.append(
                [
                    r.get("id"),
                    r.get("product_name"),
                    _parse_date(r.get("trade_date")),
                    r.get("stock_code"),
                    r.get("stock_name"),
                    float(r.get("position_pct") or 0) if r.get("position_pct") is not None else None,
                    r.get("side"),
                    float(r.get("price") or 0) if r.get("price") is not None else None,
                    _fmt_dt(r.get("created_at")),
                    _fmt_dt(r.get("updated_at")),
                ]
            )

        # Sheet2: 净值
        ws2 = wb.create_sheet("净值")
        ws2.append(["产品名称", "日期", "组合净值", "沪深300净值"])
        for r in rows2:
            ws2.append(
                [
                    name,
                    _parse_date(r.get("biz_date")),
                    float(r.get("nav") or 0) if r.get("nav") is not None else None,
                    float(r.get("hs300_nav") or 0) if r.get("hs300_nav") is not None else None,
                ]
            )

        # Sheet3: 净值明细（row_type=2）
        ws3 = wb.create_sheet("净值明细")
        ws3.append(["日期", "个股", "股票代码", "持仓份额"])
        for r in rows3:
            ws3.append(
                [
                    _parse_date(r.get("biz_date")),
                    r.get("stock_name"),
                    r.get("stock_code"),
                    float(r.get("position_after") or 0) if r.get("position_after") is not None else None,
                ]
            )

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        fname = f"{name}_仓位+净值.xlsx"
        headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"}
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# 净值图：计算参数与 temp3.py 一致
NAV_INIT_CAPITAL = 10_000_000  # 初始资金 1000万
NAV_TRADE_UNIT = 100           # A股交易单位（股）
HSGT_PRICE_TABLE = "portal_db.hsgt_price_deliver"  # 收盘价表，biz_date=yyyymmdd, stock_code, last_price
PRODUCT_NAV_DAILY_TABLE = "product_nav_daily"     # 净值结果表，由定时任务按同口径写入，接口优先查此表
# 净值数据来源：product_nav_daily（默认）或 mv_product_nav_compute（物化视图直接计算，含沪深300）
NAV_SOURCE_TABLE = os.environ.get("NAV_SOURCE_TABLE", "product_nav_daily")
PRODUCT_NAV_DAILY_DETAIL_TABLE = "product_nav_daily_detail"  # 净值明细表，row_type=3 为当日汇总行


def _parse_date(v: Any) -> Optional[str]:
    """归一化为 YYYY-MM-DD。"""
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if len(s) >= 8 and s.isdigit():
        return s[:4] + "-" + s[4:6] + "-" + s[6:8]
    if "-" in s:
        return s[:10]
    return None


def _normalize_biz_date_to_yyyymmdd(v: Any) -> Optional[str]:
    """
    hsgt_price_deliver 的 biz_date 口径在不同查询结果里可能是：
    - date/datetime 类型
    - '20260403'（字符串）
    - '2026-04-03'（字符串）
    这里统一转成 'YYYYMMDD' 以保证 WHERE biz_date = %s 能命中。
    """
    if v is None:
        return None
    if hasattr(v, "strftime"):
        return v.strftime("%Y%m%d")
    s = str(v).strip()
    if not s:
        return None
    # 去掉所有非数字字符（把 2026-04-03 变成 20260403）
    digits = _re.sub(r"[^0-9]", "", s)
    if len(digits) < 8:
        return None
    return digits[:8]


def _compute_nav_series(product_name: str, cur) -> List[Dict[str, Any]]:
    """
    按 temp3 口径计算组合净值序列。
    交易来自 config_stock_position，收盘价来自 hsgt_price_deliver（biz_date=yyyymmdd, stock_code, last_price）。
    """
    cur.execute(
        f"SELECT trade_date, stock_code, side, position_pct, price "
        f"FROM `{CONFIG_STOCK_POSITION_TABLE}` "
        f"WHERE product_name = %s AND trade_date IS NOT NULL AND stock_code IS NOT NULL AND side IS NOT NULL "
        f"ORDER BY trade_date ASC",
        (product_name,),
    )
    trades = cur.fetchall()
    if not trades:
        return []

    # 日期范围
    dates_from_trades = set()
    for r in trades:
        d = _parse_date(r.get("trade_date"))
        if d:
            dates_from_trades.add(d)

    if not dates_from_trades:
        return []
    min_d, max_d = min(dates_from_trades), max(dates_from_trades)

    # 从 hsgt_price_deliver 拉取区间内收盘价（biz_date 可能是 yyyymmdd 或 date 类型）
    cur.execute(
        f"SELECT biz_date, stock_code, last_price FROM `{HSGT_PRICE_TABLE}` "
        f"WHERE biz_date IS NOT NULL AND last_price IS NOT NULL"
    )
    price_rows = cur.fetchall()

    # 构建 (date_str, stock_code) -> price；日期统一为 YYYY-MM-DD
    price_map: Dict[tuple, float] = {}
    trade_dates_set = set()
    for r in price_rows:
        dt = _parse_date(r.get("biz_date"))
        if not dt:
            continue
        code = (r.get("stock_code") or "").strip()
        if not code:
            continue
        try:
            pr = float(r.get("last_price"))
        except (TypeError, ValueError):
            continue
        price_map[(dt, code)] = pr
        trade_dates_set.add(dt)

    # 交易日历：交易日期与有价格数据的日期的并集，且在 [min_d, max_d] 内
    all_dates = dates_from_trades | trade_dates_set
    sorted_dates = sorted(d for d in all_dates if min_d <= d <= max_d)
    if not sorted_dates:
        return []

    # 按日期分组的交易
    trade_by_date: Dict[str, list] = {}
    for r in trades:
        d = _parse_date(r.get("trade_date"))
        if d:
            trade_by_date.setdefault(d, []).append(r)

    def _get_price(date_str: str, code: str) -> float:
        v = price_map.get((date_str, code))
        if v is not None:
            return v
        # 尝试 SZ/SH 前缀对齐
        if code.upper().startswith(("SH", "SZ")) and len(code) > 2:
            v = price_map.get((date_str, code[2:]))
            if v is not None:
                return v
        if len(code) == 6 or len(code) <= 6:
            for p in ("SH", "SZ"):
                v = price_map.get((date_str, p + code))
                if v is not None:
                    return v
        return 0.0

    cash = float(NAV_INIT_CAPITAL)
    positions: Dict[str, int] = {}
    result: List[Dict[str, Any]] = []

    for date_str in sorted_dates:
        day_trades = trade_by_date.get(date_str, [])
        for row in day_trades:
            code = (row.get("stock_code") or "").strip()
            side = (row.get("side") or "").strip()
            try:
                pct = float(row.get("position_pct") or 0)
                price = float(row.get("price") or 0)
            except (TypeError, ValueError):
                continue
            if price <= 0:
                continue
            target_amount = NAV_INIT_CAPITAL * (pct / 100.0)

            if side == "买入":
                shares = int(target_amount / price / NAV_TRADE_UNIT) * NAV_TRADE_UNIT
                if shares <= 0:
                    continue
                cost = shares * price
                if cash < cost:
                    shares = int(cash / price / NAV_TRADE_UNIT) * NAV_TRADE_UNIT
                    if shares <= 0:
                        continue
                    cost = shares * price
                cash -= cost
                positions[code] = positions.get(code, 0) + shares
            elif side == "卖出":
                if code not in positions or positions[code] <= 0:
                    continue
                shares = int(target_amount / price / NAV_TRADE_UNIT) * NAV_TRADE_UNIT
                shares = min(shares, positions[code])
                if shares <= 0:
                    continue
                cash += shares * price
                positions[code] -= shares
                if positions[code] == 0:
                    del positions[code]

        market_value = 0.0
        for c, sh in positions.items():
            market_value += sh * _get_price(date_str, c)
        total_asset = cash + market_value
        nav = total_asset / float(NAV_INIT_CAPITAL)
        result.append({"date": date_str, "nav": round(nav, 6), "hs300_nav": None})

    return result


class NavChartPoint(BaseModel):
    date: str
    nav: float
    hs300_nav: Optional[float] = None
    # 当日组合总仓位（来自 product_nav_daily_detail row_type=3.open_pct，百分比数值）
    open_pct: Optional[float] = None


class NavDetailRow(BaseModel):
    biz_date: str
    stock_name: Optional[str] = None
    stock_code: Optional[str] = None
    # 仓位：等价于 product_nav_daily_detail.open_pct（百分比数值，如 5 表示 5%）
    open_pct: Optional[float] = None
    position_after: Optional[float] = None


class NavDetailResp(BaseModel):
    product_name: str
    biz_date: Optional[str] = None
    total: int = 0
    items: List[NavDetailRow] = []


@app.get("/api/config/stock-position/nav-chart", response_model=List[NavChartPoint])
async def get_stock_position_nav_chart(
    product_name: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> List[NavChartPoint]:
    """
    返回指定产品的净值序列，用于前端绘制净值图。
    直接从 product_nav_daily_detail 中查询 row_type=3 的当日汇总行（便于对账，且无需实时计算）。

    - 默认返回最近 90 天（约 3 个月）
    - 可通过 start_date/end_date（YYYY-MM-DD）指定区间
    """
    name = (product_name or "").strip() or "短线王"
    try:
        with _db_cursor() as cur:
            # 默认区间：最近 90 天
            from datetime import date, timedelta
            d_to = end_date.strip() if end_date and end_date.strip() else date.today().strftime("%Y-%m-%d")
            if start_date and start_date.strip():
                d_from = start_date.strip()
            else:
                d_from = (date.today() - timedelta(days=90)).strftime("%Y-%m-%d")

            cur.execute(
                f"SELECT biz_date, nav, hs300_nav, open_pct FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                f"WHERE product_name = %s AND row_type = 3 "
                f"AND biz_date BETWEEN %s AND %s "
                f"ORDER BY biz_date ASC",
                (name, d_from, d_to),
            )
            rows = cur.fetchall()

            # 基期锚点必须按“该产品首个交易日的上一个交易日”确定，不能用统一日期。
            # 这里先取该产品首个 trade_date，再从净值结果表取其之前最近一个点作为锚点。
            anchor_row = None
            cur.execute(
                f"SELECT MIN(trade_date) AS min_trade_date "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE product_name = %s AND trade_date IS NOT NULL",
                (name,),
            )
            tr = cur.fetchone() or {}
            min_trade_date = tr.get("min_trade_date")
            if min_trade_date:
                cur.execute(
                    f"SELECT biz_date, nav, hs300_nav, open_pct FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                    f"WHERE product_name = %s AND row_type = 3 "
                    f"AND biz_date < %s "
                    f"ORDER BY biz_date DESC LIMIT 1",
                    (name, min_trade_date),
                )
                anchor_row = cur.fetchone() or None

            # 防止“锚点日期”落在当前净值区间内部，导致同一 biz_date（例如 2026-03-30）在
            # 开头/结尾各出现一次，从而在图表和列表中看到重复日期。
            if rows and anchor_row:
                try:
                    first_row_d = _parse_date(rows[0].get("biz_date"))
                    last_row_d = _parse_date(rows[-1].get("biz_date"))
                    anchor_d = _parse_date(anchor_row.get("biz_date"))
                except Exception:
                    first_row_d = last_row_d = anchor_d = None

                if (
                    anchor_d is None
                    or (first_row_d is not None and last_row_d is not None and first_row_d <= anchor_d <= last_row_d)
                ):
                    # 若锚点日期已在当前区间内（或解析失败），则丢弃锚点，避免重复点
                    anchor_row = None
        series = []
        if anchor_row:
            dt0 = _parse_date(anchor_row.get("biz_date"))
            if dt0:
                try:
                    nav0 = float(anchor_row.get("nav") or 0)
                    hs0_raw = anchor_row.get("hs300_nav")
                    hs0 = float(hs0_raw) if hs0_raw is not None else None
                    op0_raw = anchor_row.get("open_pct")
                    op0 = float(op0_raw) if op0_raw is not None else None
                    series.append({"date": dt0, "nav": round(nav0, 6), "hs300_nav": hs0, "open_pct": op0})
                except (TypeError, ValueError):
                    pass
        for r in rows:
            dt = _parse_date(r.get("biz_date"))
            if not dt:
                continue
            try:
                nav_val = float(r.get("nav") or 0)
            except (TypeError, ValueError):
                continue
            hs = r.get("hs300_nav")
            try:
                hs300_val = float(hs) if hs is not None else None
            except (TypeError, ValueError):
                hs300_val = None
            op = r.get("open_pct")
            try:
                open_pct_val = float(op) if op is not None else None
            except (TypeError, ValueError):
                open_pct_val = None
            if series and series[-1].get("date") == dt:
                # 避免补点与区间首点同日重复
                continue
            series.append({"date": dt, "nav": round(nav_val, 6), "hs300_nav": hs300_val, "open_pct": open_pct_val})

        # 只有 1 天数据时，前端会判定“<2点不可绘制”。这里补一个“初始点=1”以保证可画图。
        # 说明：补点日期取首日的前一自然日；不影响多日序列的正常展示。
        if len(series) == 1:
            try:
                from datetime import datetime, timedelta

                first_d = datetime.strptime(series[0]["date"], "%Y-%m-%d").date()
                prev_d = first_d - timedelta(days=1)
                series.insert(0, {"date": prev_d.strftime("%Y-%m-%d"), "nav": 1.0, "hs300_nav": 1.0, "open_pct": None})
            except Exception:
                pass
        return [NavChartPoint(**x) for x in series]
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/stock-position/nav-detail", response_model=NavDetailResp)
async def get_stock_position_nav_detail(
    product_name: Optional[str] = None,
    biz_date: Optional[str] = None,
) -> NavDetailResp:
    """
    返回“净值明细”（product_nav_daily_detail row_type=2）
    - 默认：仅返回最新 biz_date 的明细
    - 过滤：按 product_name（即配置页选择的产品）
    - 字段：biz_date, stock_name, stock_code, open_pct, position_after
    """
    name = (product_name or "").strip() or "短线王"
    try:
        with _db_cursor() as cur:
            if biz_date and biz_date.strip():
                target_biz_date = biz_date.strip()
            else:
                # 取最新 biz_date（order by + limit，兼容 biz_date 是 date 或字符串）
                cur.execute(
                    f"SELECT biz_date FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                    f"WHERE product_name = %s AND row_type = 2 AND biz_date IS NOT NULL "
                    f"ORDER BY biz_date DESC LIMIT 1",
                    (name,),
                )
                r = cur.fetchone() or {}
                target_biz_date = r.get("biz_date")

            dt_norm = _parse_date(target_biz_date)

            # 查指定（或最新）biz_date 的明细
            cur.execute(
                f"SELECT biz_date, stock_name, stock_code, open_pct, position_after "
                f"FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                f"WHERE product_name = %s AND row_type = 2 "
                f"  AND biz_date IS NOT NULL "
                f"  AND TRIM(CAST(biz_date AS STRING)) != '' "
                f"  AND biz_date = %s "
                f"ORDER BY stock_code ASC",
                (name, target_biz_date),
            )
            rows = cur.fetchall() or []

        items: List[NavDetailRow] = []
        for r in rows:
            items.append(
                NavDetailRow(
                    biz_date=_parse_date(r.get("biz_date")) or str(r.get("biz_date") or ""),
                    stock_name=r.get("stock_name"),
                    stock_code=r.get("stock_code"),
                    open_pct=float(r.get("open_pct") or 0) if r.get("open_pct") is not None else None,
                    position_after=float(r.get("position_after") or 0) if r.get("position_after") is not None else None,
                )
            )

        # 追加“总计”汇总行：
        # - 日期列：总计
        # - 个股列：N只
        # - 股票代码列：空
        # - 仓位列：sum(open_pct)
        # - 持仓份额列：sum(position_after)
        if items:
            total_cnt = len(items)
            total_open_pct = 0.0
            total_position_after = 0.0
            has_open_pct = False
            has_position_after = False
            for it in items:
                if it.open_pct is not None:
                    has_open_pct = True
                    total_open_pct += float(it.open_pct or 0)
                if it.position_after is not None:
                    has_position_after = True
                    total_position_after += float(it.position_after or 0)
            items.append(
                NavDetailRow(
                    biz_date="总计",
                    stock_name=f"{total_cnt}只",
                    stock_code="",
                    open_pct=total_open_pct if has_open_pct else None,
                    position_after=total_position_after if has_position_after else None,
                )
            )
        return NavDetailResp(
            product_name=name,
            biz_date=dt_norm,
            total=len(items),
            items=items,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/config/stock-position/nav-export.xlsx")
async def export_stock_position_nav_excel(
    product_name: Optional[str] = None,
):
    """
    下载指定产品的净值（来自 product_nav_daily_detail row_type=3），全量导出为 Excel。
    """
    name = (product_name or "").strip() or "短线王"
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT biz_date, total_position_mv, total_cash, total_asset, nav, hs300_nav "
                f"FROM `{PRODUCT_NAV_DAILY_DETAIL_TABLE}` "
                f"WHERE product_name = %s AND row_type = 3 "
                f"ORDER BY biz_date ASC",
                (name,),
            )
            rows = cur.fetchall()

        from openpyxl import Workbook

        wb = Workbook()
        ws = wb.active
        ws.title = "净值"
        ws.append(["日期", "持仓市值", "现金", "总资产", "组合净值", "沪深300净值"])

        for r in rows:
            dt = _parse_date(r.get("biz_date")) or ""
            ws.append(
                [
                    dt,
                    float(r.get("total_position_mv") or 0),
                    float(r.get("total_cash") or 0),
                    float(r.get("total_asset") or 0),
                    float(r.get("nav") or 0),
                    float(r.get("hs300_nav") or 0) if r.get("hs300_nav") is not None else None,
                ]
            )

        bio = io.BytesIO()
        wb.save(bio)
        bio.seek(0)

        fname = f"{name}_净值.xlsx"
        headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname)}"}
        return StreamingResponse(
            bio,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/config/stock-position", response_model=StockPositionOut)
async def create_stock_position(body: StockPositionCreate, request: Request) -> StockPositionOut:
    _reject_if_readonly(request)
    """新增：股票代码、仓位、买入/卖出、成交价为必填；日期不填则默认当前日期。"""
    from datetime import date
    trade_date = body.trade_date
    if not trade_date or not trade_date.strip():
        trade_date = date.today().strftime("%Y-%m-%d")
    try:
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            product_name = (body.product_name or "").strip() or None
            next_row_id = _next_row_id_for_date(cur, product_name or "", trade_date.strip())
            next_id = _next_config_id(cur, CONFIG_STOCK_POSITION_TABLE)
            if not next_id or next_id < 1:
                # 理论上不该发生；兜底避免再次写入 NULL/非法 id
                next_id = 1

            # -------------------- 填报校验：stock_name 必须与价格表一致 --------------------
            # 规则：在 hsgt_price_deliver 取 biz_date 最大的一天；用 stock_code 关联取 stock_name；
            # 如果与页面填的 stock_name 不一致，直接拦截。
            stock_code_input = (body.stock_code or "").strip()
            stock_name_input = (body.stock_name or "").strip() if body.stock_name else ""
            if stock_code_input and stock_name_input:
                # 严格按 stock_code 关联 hsgt_price_deliver（不做前缀映射）
                candidate_codes: List[str] = [stock_code_input]

                cur.execute(
                    "SELECT MAX(biz_date) AS max_d FROM portal_db.hsgt_price_deliver WHERE biz_date IS NOT NULL"
                )
                max_d_row = cur.fetchone() or {}
                max_d = max_d_row.get("max_d")
                max_d_norm = _normalize_biz_date_to_yyyymmdd(max_d)
                if max_d_norm is not None:
                    placeholders = ",".join(["%s"] * len(candidate_codes))
                    cur.execute(
                        f"""
                        SELECT stock_name
                        FROM portal_db.hsgt_price_deliver
                        WHERE biz_date = %s
                          AND stock_code IN ({placeholders})
                          AND stock_name IS NOT NULL
                          AND TRIM(CAST(stock_name AS STRING)) != ''
                        LIMIT 1
                        """,
                        (max_d_norm, *candidate_codes),
                    )
                    expected_row = cur.fetchone() or {}
                    expected_name = (expected_row.get("stock_name") or "").strip()
                    if not expected_name or expected_name != stock_name_input:
                        raise HTTPException(status_code=400, detail="填报数校验不对")
            cur.execute(
                f"INSERT INTO `{CONFIG_STOCK_POSITION_TABLE}` "
                f"(id, product_name, trade_date, row_id, stock_code, stock_name, position_pct, side, price, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (
                    next_id,
                    product_name,
                    trade_date.strip(),
                    next_row_id,
                    (body.stock_code or "").strip(),
                    (body.stock_name or "").strip() or None,
                    body.position_pct,
                    (body.side or "").strip(),
                    body.price,
                    now_str,
                    now_str,
                ),
            )

            cur.execute(
                f"SELECT id, product_name, trade_date, row_id, stock_code, stock_name, "
                f"position_pct, side, price, created_at, updated_at "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` "
                f"WHERE product_name <=> %s AND trade_date = %s AND row_id = %s "
                f"AND id = %s "
                f"ORDER BY id DESC LIMIT 1",
                (product_name or None, trade_date.strip(), next_row_id, next_id),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=500, detail="创建失败")
        return StockPositionOut(**_row_stock_position_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.put("/api/config/stock-position/{item_id}", response_model=StockPositionOut)
async def update_stock_position(item_id: int, body: StockPositionUpdate, request: Request) -> StockPositionOut:
    _reject_if_readonly(request)
    """全字段可选更新；未传的字段保持原值。"""
    try:
        from datetime import date
        now_str = datetime.now().strftime(_DT_FMT)
        with _db_cursor() as cur:
            cur.execute(
                f"SELECT id, product_name, trade_date, row_id, stock_code, stock_name, "
                f"position_pct, side, price, created_at, updated_at "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` WHERE id = %s",
                (item_id,),
            )
            old = cur.fetchone()
            if not old:
                raise HTTPException(status_code=404, detail="记录不存在")

            def _v(key: str, new_val: Any, default: Any = None):
                if new_val is not None:
                    return new_val
                return old.get(key) if old else default

            trade_date = _v("trade_date", body.trade_date)
            if trade_date and isinstance(trade_date, str) and not trade_date.strip():
                trade_date = date.today().strftime("%Y-%m-%d")

            new_product = _v("product_name", body.product_name)
            new_product = (new_product or "").strip() or None
            new_row_id = _v("row_id", None) or old.get("row_id")
            new_code = (_v("stock_code", body.stock_code) or "").strip()
            new_name = (_v("stock_name", body.stock_name) or "").strip() or None
            new_pct = _v("position_pct", body.position_pct)
            new_side = (_v("side", body.side) or "").strip()
            new_price = _v("price", body.price)
            created_at = _fmt_dt(old.get("created_at")) or now_str

            # -------------------- 填报校验：stock_name 必须与价格表一致 --------------------
            # 规则：hsgt_price_deliver 取 biz_date 最大的一天，然后按 stock_code 查 stock_name 比较。
            # 只有当用户显式提供 new_name 时才校验（避免允许 stock_name 留空）。
            if new_code and new_name:
                # 严格按 stock_code 关联 hsgt_price_deliver（不做前缀映射）
                candidate_codes: List[str] = [new_code]

                cur.execute(
                    "SELECT MAX(biz_date) AS max_d FROM portal_db.hsgt_price_deliver WHERE biz_date IS NOT NULL"
                )
                max_d_row = cur.fetchone() or {}
                max_d = max_d_row.get("max_d")
                max_d_norm = _normalize_biz_date_to_yyyymmdd(max_d)
                if max_d_norm is not None:
                    placeholders = ",".join(["%s"] * len(candidate_codes))
                    cur.execute(
                        f"""
                        SELECT stock_name
                        FROM portal_db.hsgt_price_deliver
                        WHERE biz_date = %s
                          AND stock_code IN ({placeholders})
                          AND stock_name IS NOT NULL
                          AND TRIM(CAST(stock_name AS STRING)) != ''
                        LIMIT 1
                        """,
                        (max_d_norm, *candidate_codes),
                    )
                    expected_row = cur.fetchone() or {}
                    expected_name = (expected_row.get("stock_name") or "").strip()
                    if not expected_name or expected_name != new_name:
                        raise HTTPException(status_code=400, detail="填报数校验不对")

            cur.execute(
                f"DELETE FROM `{CONFIG_STOCK_POSITION_TABLE}` WHERE id = %s",
                (item_id,),
            )
            cur.execute(
                f"INSERT INTO `{CONFIG_STOCK_POSITION_TABLE}` "
                f"(id, product_name, trade_date, row_id, stock_code, stock_name, position_pct, side, price, created_at, updated_at) "
                f"VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)",
                (
                    item_id,
                    new_product,
                    trade_date or date.today().strftime("%Y-%m-%d"),
                    new_row_id,
                    new_code,
                    new_name,
                    new_pct,
                    new_side,
                    new_price,
                    created_at,
                    now_str,
                ),
            )
            cur.execute(
                f"SELECT id, product_name, trade_date, stock_code, stock_name, "
                f"position_pct, side, price, created_at, updated_at "
                f"FROM `{CONFIG_STOCK_POSITION_TABLE}` WHERE id = %s",
                (item_id,),
            )
            row = cur.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="记录不存在")
        return StockPositionOut(**_row_stock_position_fix(row))
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.delete("/api/config/stock-position/{item_id}")
async def delete_stock_position(item_id: int, request: Request) -> Dict[str, Any]:
    _reject_if_readonly(request)
    try:
        with _db_cursor() as cur:
            cur.execute(
                f"DELETE FROM `{CONFIG_STOCK_POSITION_TABLE}` WHERE id = %s",
                (item_id,),
            )
            if cur.rowcount == 0:
                raise HTTPException(status_code=404, detail="记录不存在")
        return {"success": True}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# -------------------- API：沪深港通价格推送（portal_db.hsgt_price_deliver） --------------------

class HsgtPriceDeliverOut(BaseModel):
    biz_date: Optional[str] = None  # YYYY-MM-DD
    stock_code: Optional[str] = None
    stock_name: Optional[str] = None
    last_price: Optional[float] = None


@app.get("/api/hsgt-price-deliver", response_model=List[HsgtPriceDeliverOut])
async def get_hsgt_price_deliver(
    biz_date: Optional[str] = None,
) -> List[HsgtPriceDeliverOut]:
    """
    查询 StarRocks 中 portal_db.hsgt_price_deliver 表。

    - 若传入 biz_date（格式如 '2026-02-11'），则按该日期筛选；
    - 若不传 biz_date，则返回表中全部数据。
    """
    try:
        with _db_cursor() as cur:
            sql = (
                "SELECT biz_date, stock_code, stock_name, last_price "
                "FROM portal_db.hsgt_price_deliver where biz_date is not null"
            )
            params: List[Any] = []
            if biz_date:
                sql += " and biz_date = %s"
                params.append(biz_date)
            cur.execute(sql, tuple(params))
            rows = cur.fetchall()
        # rows 已是 DictCursor 结果，字段名与模型一致，直接解包
        return [HsgtPriceDeliverOut(**row) for row in rows]
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

if FRONTEND_DIST.exists():
    _assets_dir = FRONTEND_DIST / "assets"

    @app.get("/assets/{file_path:path}")
    async def serve_assets(file_path: str):
        """显式路由优先于 catch-all；带 hash 的 JS/CSS 强缓存，来回切换时从缓存加载"""
        if ".." in file_path or file_path.startswith("/"):
            raise HTTPException(status_code=404, detail="Not Found")
        full = _assets_dir / file_path
        if not full.is_file():
            raise HTTPException(status_code=404, detail="Not Found")
        # 文件名带 hash（如 index-Bs3PAl4u.js），可长期强缓存；来回切换业务模块时不再重复下载
        headers = {"Cache-Control": "public, max-age=31536000, immutable"}
        return FileResponse(full, headers=headers)

    def _index_html(base_path: Optional[str] = None):
        """base_path: 门户代理时传 /sr_api，使静态资源与 API 请求走同源"""
        path = (base_path or MOUNT_PATH or "").strip("/")
        mount = ("/" + path) if path else ""
        html = (FRONTEND_DIST / "index.html").read_text(encoding="utf-8")
        if mount:
            html = html.replace('src="/assets/', 'src="' + mount + '/assets/').replace('href="/assets/', 'href="' + mount + '/assets/')
            if mount != "/sr_api":
                html = html.replace('src="/sr_api/', 'src="' + mount + '/').replace('href="/sr_api/', 'href="' + mount + '/')
            inject = '<script>window.__API_BASE__="' + mount + '";</script>'
            html = html.replace("<head>", "<head>" + inject, 1) if "<head>" in html else html.replace("<body>", "<body>" + inject, 1)
        return html

    _no_cache_headers = {"Cache-Control": "no-store, no-cache, must-revalidate", "Pragma": "no-cache"}

    @app.get("/")
    async def index(req: Request):
        prefix = (req.headers.get("X-Forwarded-Prefix") or "").strip()
        return HTMLResponse(_index_html(prefix or None), headers=_no_cache_headers)
    @app.get("/{path:path}")
    async def serve_spa(req: Request, path: str):
        if path.strip("/").startswith("assets"):
            raise HTTPException(status_code=404, detail="Not Found")
        prefix = (req.headers.get("X-Forwarded-Prefix") or "").strip()
        return HTMLResponse(_index_html(prefix or None), headers=_no_cache_headers)
else:
    @app.get("/")
    async def index():
        return {"status": "ok", "tip": "请先执行 cd business/frontend && VITE_BASE=/business/realtime_mv/ npm run build 构建前端"}

if __name__ == "__main__":
    port = int(os.environ.get("BUSINESS_PORT", "5003"))
    import uvicorn
    print("业务应用（端口 %s）: http://127.0.0.1:%s" % (port, port))
    uvicorn.run(app, host="0.0.0.0", port=port)
